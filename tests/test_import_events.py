"""Tests for the import feature (v2: events + users + dynamic spots + assignments)."""

import importlib.util
import json
import sys
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
import pytest

from app.extensions import db
from app.models.assignment import DebriefingRecord
from app.models.audit import AuditLogEntry
from app.models.event import Event, EventStatus
from app.models.qualification import Qualification
from app.models.role import Role
from app.models.user import UserAccount
from tests.conftest import _get_csrf, _make_master_event
from tests.conftest import _make_user as _conftest_make_user

# ── Load the extraction script without adding it to the package ────────────────


def _import_script():
    script_path = Path(__file__).parent.parent / "scripts" / "import_events.py"
    spec = importlib.util.spec_from_file_location("import_script", script_path)
    assert spec and spec.loader
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore[union-attr]
    return mod


_script = _import_script()

# ── Script unit tests (no DB) ─────────────────────────────────────────────────


class TestReverseNameHelper:
    def test_two_part_name(self):
        assert _script._reverse_name("Balhar Lumír") == "Lumír Balhar"

    def test_three_part_name(self):
        assert _script._reverse_name("Svobodová K. Zuzana") == "K. Zuzana Svobodová"

    def test_single_part_unchanged(self):
        assert _script._reverse_name("Novák") == "Novák"

    def test_strips_whitespace(self):
        assert _script._reverse_name("  Gajda  Adam  ") == "Adam Gajda"


class TestIsValidNameHelper:
    def test_valid_names(self):
        assert _script._is_valid_name("Adam Gajda")
        assert _script._is_valid_name("X")

    def test_junk_strings(self):
        assert not _script._is_valid_name(".")
        assert not _script._is_valid_name("123")
        assert not _script._is_valid_name("")


class TestFmtTimeHelper:
    """Unit tests for _fmt_time()."""

    def test_normal_time(self):

        assert _script._fmt_time(time(10, 30)) == "10:30"

    def test_midnight(self):

        assert _script._fmt_time(time(0, 0)) == "00:00"

    def test_none_returns_none(self):
        assert _script._fmt_time(None) is None

    def test_non_time_returns_none(self):
        assert _script._fmt_time("10:30") is None
        assert _script._fmt_time(1030) is None


class TestFmtDateHelper:
    """Unit tests for _fmt_date()."""

    def test_datetime_object(self):

        assert _script._fmt_date(datetime(2026, 5, 15)) == "2026-05-15"

    def test_none_returns_none(self):
        assert _script._fmt_date(None) is None

    def test_string_returns_none(self):
        assert _script._fmt_date("2026-05-15") is None

    def test_date_object_returns_none(self):

        # _fmt_date expects datetime, not date
        assert _script._fmt_date(date(2026, 5, 15)) is None


class TestBuildDescription:
    """Unit tests for _build_description()."""

    def test_all_fields(self):
        result = _script._build_description(
            vehicle="Sanitka",
            event_type="zdravotní dozor",
            contact="Jan Novák 123",
            signups=["Petr", "Marie"],
            time_missing=False,
        )
        assert "Typ: zdravotní dozor" in result
        assert "Vozidlo/stan: Sanitka" in result
        assert "Kontakt pořadatel: Jan Novák 123" in result
        assert "Přihlášení (import z GS): Petr, Marie" in result

    def test_time_missing_warning(self):
        result = _script._build_description(
            vehicle=None,
            event_type=None,
            contact=None,
            signups=[],
            time_missing=True,
        )
        assert "UPOZORNĚNÍ" in result
        assert "Čas akce" in result

    def test_empty_inputs(self):
        result = _script._build_description(vehicle=None, event_type=None, contact=None, signups=[], time_missing=False)
        assert result == ""

    def test_signups_with_empty_strings_filtered(self):
        result = _script._build_description(
            vehicle=None, event_type=None, contact=None, signups=["Jan", "", "  "], time_missing=False
        )
        assert "Jan" in result
        assert result.count(",") == 0  # only one valid name


class TestIsRowCancelled:
    """Unit tests for _is_row_cancelled()."""

    def test_normal_row_not_cancelled(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture))
        ws = wb["Dozory"]
        # Row 3 is a normal event
        row = list(ws.iter_rows(min_row=3, max_row=3))[0]
        assert _script._is_row_cancelled(row) is False


class TestExtractFunction:
    """Unit tests for extract() using the test fixture."""

    def test_extracts_all_events(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        events = _script.extract(wb)
        assert len(events) == 9

    def test_event_basic_fields(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        events = _script.extract(wb)
        first = events[0]
        assert first["name"] == "Sportovní závody"
        assert first["date"] == "2026-07-15"
        assert first["start_time"] == "10:00"
        assert first["end_time"] == "14:00"
        assert first["location"] == "Sportovní hala Testov"
        assert first["paid"] is False
        assert first["cancelled"] is False
        assert first["time_missing"] is False

    def test_paid_event(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        events = _script.extract(wb)
        # Hasičský ples (row 4) is paid=True
        ples = next(e for e in events if "Hasičský ples" in e["name"])
        assert ples["paid"] is True

    def test_time_missing_event(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        events = _script.extract(wb)
        # Kulturní festival (row 6) has start_time=None
        festival = next(e for e in events if "Kulturní festival" in e["name"])
        assert festival["time_missing"] is True
        assert festival["start_time"] is None
        assert "UPOZORNĚNÍ" in festival["description"]

    def test_duplicate_names_get_date_suffix(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        events = _script.extract(wb)
        # Fotbalový turnaj appears twice (rows 7, 8) — should get date suffix
        fotbal = [e for e in events if "Fotbalový turnaj" in e["name"]]
        assert len(fotbal) == 2
        assert fotbal[0]["name"] != fotbal[1]["name"]
        # Should have date suffixes
        assert "1.10." in fotbal[0]["name"] or "15.10." in fotbal[0]["name"]

    def test_midnight_end_time_treated_as_null(self):
        """Issue #340: midnight (00:00) end time is treated as unspecified."""
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        events = _script.extract(wb)
        # Noční akce (row 11) has end_time=time(0,0)
        nocni = next(e for e in events if "Noční akce" in e["name"])
        assert nocni["start_time"] == "20:00"
        # Current code converts midnight to None — this tests the CURRENT behavior
        assert nocni["end_time"] is None

    def test_cutoff_filters_events(self):

        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        # Only events on or after 2026-10-01
        events = _script.extract(wb, cutoff=date(2026, 10, 1))
        dates = [e["date"] for e in events]
        assert all(d >= "2026-10-01" for d in dates)
        # Should exclude July, August, September events
        assert not any(d.startswith("2026-07") for d in dates)

    def test_responsible_person_extracted(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        events = _script.extract(wb)
        first = events[0]
        assert first["responsible_person"] == "Novák Jan"

    def test_signups_extracted(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        events = _script.extract(wb)
        # Hasičský ples has multiple signups (cols N+)
        ples = next(e for e in events if "Hasičský ples" in e["name"])
        assert len(ples["signups"]) >= 2

    def test_description_includes_event_type(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        events = _script.extract(wb)
        first = events[0]
        assert "zdravotní dozor" in first["description"]


class TestExtractUsersFunction:
    """Unit tests for extract_users() using the test fixture."""

    def test_extracts_users(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        users = _script.extract_users(wb)
        assert len(users) > 0
        # Users should be sorted by name
        names = [u["name"] for u in users]
        assert names == sorted(names)

    def test_user_has_expected_fields(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        users = _script.extract_users(wb)
        for u in users:
            assert "gs_name" in u
            assert "name" in u
            assert "email" in u
            assert "phone" in u
            assert "is_zdravotnik" in u
            assert "is_ridic" in u

    def test_cutoff_filters_users(self):

        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        all_users = _script.extract_users(wb)
        filtered_users = _script.extract_users(wb, cutoff=date(2026, 10, 1))
        # Filtered should be a subset
        assert len(filtered_users) <= len(all_users)


class TestLoadLidiLookup:
    """Unit tests for _load_lidi_lookup()."""

    def test_loads_lidi_data(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        lookup = _script._load_lidi_lookup(wb)
        assert len(lookup) > 0
        # Check a known entry
        assert "Kratochvíl Tomáš" in lookup
        info = lookup["Kratochvíl Tomáš"]
        assert info["is_zdravotnik"] is True
        assert info["is_ridic"] is True

    def test_ridic_false_when_not_set(self):
        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        lookup = _script._load_lidi_lookup(wb)
        # Svoboda Petr: zdravotník=False, ridic=False
        assert "Svoboda Petr" in lookup
        assert lookup["Svoboda Petr"]["is_ridic"] is False


class TestMainCli:
    """Unit tests for main() CLI entry point."""

    def test_main_file_not_found(self, capsys):

        with pytest.raises(SystemExit) as exc_info:
            sys.argv = ["import_events.py", "--input", "/nonexistent/file.xlsx"]
            _script.main()
        assert exc_info.value.code == 1
        captured = capsys.readouterr()
        assert "ERROR" in captured.err

    def test_main_invalid_cutoff(self, capsys):

        with pytest.raises(SystemExit) as exc_info:
            sys.argv = ["import_events.py", "--input", "x.xlsx", "--cutoff", "not-a-date"]
            _script.main()
        assert exc_info.value.code == 1
        captured = capsys.readouterr()
        assert "Invalid --cutoff" in captured.err

    def test_main_stdout_output(self, capsys, tmp_path):

        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        sys.argv = ["import_events.py", "--input", str(fixture), "--output", "-"]
        _script.main()
        captured = capsys.readouterr()
        data = json.loads(captured.out)
        assert data["version"] == 2
        assert "users" in data
        assert "events" in data

    def test_main_file_output(self, tmp_path):

        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        out = tmp_path / "out.json"
        sys.argv = ["import_events.py", "--input", str(fixture), "--output", str(out)]
        _script.main()
        assert out.exists()
        data = json.loads(out.read_text())
        assert data["version"] == 2
        assert len(data["events"]) == 9

    def test_main_with_cutoff(self, tmp_path):

        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        out = tmp_path / "out.json"
        sys.argv = [
            "import_events.py",
            "--input",
            str(fixture),
            "--output",
            str(out),
            "--cutoff",
            "2026-10-01",
        ]
        _script.main()
        data = json.loads(out.read_text())
        assert all(e["date"] >= "2026-10-01" for e in data["events"])


# ── Route helpers ─────────────────────────────────────────────────────────────


def _make_user(app, name: str, email: str, is_zdravotnik: bool = False) -> str:
    """Create an active Member user and return its UUID string."""
    with app.app_context():
        u = _conftest_make_user(email, name, Role.MEMBER)
        return str(u.id)


def _minimal_event(name: str = "Test akce", date: str = "2030-05-01") -> dict:
    return {
        "name": name,
        "date": date,
        "start_time": "10:00",
        "end_time": "12:00",
        "location": None,
        "paid": False,
        "responsible_person": None,
        "contact_person": None,
        "description": "",
        "time_missing": False,
        "cancelled": False,
        "signups": [],
    }


def _post_confirm(
    app,
    admin_client,
    events: list[dict],
    users: list[dict] | None = None,
    master_event_id: int | None = None,
    zdravotnik_qual_id: int | None = None,
    zelenac_qual_id: int | None = None,
):
    """Build and POST the import confirm form; return the Flask response."""
    if users is None:
        users = []
    csrf = _get_csrf(admin_client, "/import/events/")

    data: dict[str, str] = {
        "csrf_token": csrf,
        "event_count": str(len(events)),
        "user_count": str(len(users)),
        "global_master_event_id": str(master_event_id or ""),
        "global_zdravotnik_qual_id": str(zdravotnik_qual_id or ""),
        "global_zelenac_qual_id": str(zelenac_qual_id or ""),
    }

    for i, ev in enumerate(events):
        p = f"ev_{i}_"
        data[f"{p}include"] = "1"
        data[f"{p}name"] = ev.get("name", "")
        data[f"{p}date"] = ev.get("date", "2030-05-01")
        data[f"{p}start_time"] = ev.get("start_time", "10:00")
        data[f"{p}end_time"] = ev.get("end_time", "12:00")
        data[f"{p}location"] = ev.get("location") or ""
        data[f"{p}paid"] = "1" if ev.get("paid") else ""
        data[f"{p}contact_person"] = ev.get("contact_person") or ""
        data[f"{p}description"] = ev.get("description") or ""
        data[f"{p}time_missing"] = "1" if ev.get("time_missing") else "0"
        data[f"{p}cancelled"] = "1" if ev.get("cancelled") else "0"
        data[f"{p}responsible_person_id"] = ev.get("responsible_person_id") or ""
        signups = ev.get("signups", [])
        data[f"{p}signup_count"] = str(len(signups))
        for j, sn in enumerate(signups):
            data[f"{p}signup_{j}"] = sn

    for i, u in enumerate(users):
        p = f"user_{i}_"
        data[f"{p}include"] = "1" if u.get("include", True) else ""
        data[f"{p}db_id"] = u.get("db_id") or ""
        data[f"{p}gs_name"] = u.get("gs_name") or ""
        data[f"{p}name"] = u.get("name") or ""
        data[f"{p}email"] = u.get("email") or ""
        data[f"{p}phone"] = u.get("phone") or ""
        data[f"{p}is_zdravotnik"] = "1" if u.get("is_zdravotnik") else "0"
        data[f"{p}is_ridic"] = "1" if u.get("is_ridic") else "0"

    return admin_client.post("/import/events/confirm", data=data, follow_redirects=False)


# ── Paste page tests ──────────────────────────────────────────────────────────


class TestImportPastePage:
    def test_requires_login(self, client):
        resp = client.get("/import/events/", follow_redirects=False)
        assert resp.status_code == 302

    def test_member_gets_403(self, member_client):
        resp = member_client.get("/import/events/")
        assert resp.status_code == 403

    def test_admin_can_access(self, admin_client):
        resp = admin_client.get("/import/events/")
        assert resp.status_code == 200


# ── Preview tests ─────────────────────────────────────────────────────────────


class TestImportPreview:
    def test_accepts_v1_flat_list(self, app, admin_client):
        _make_master_event(app)
        csrf = _get_csrf(admin_client, "/import/events/")
        payload = [_minimal_event()]
        resp = admin_client.post(
            "/import/events/preview",
            data={"json_data": json.dumps(payload), "csrf_token": csrf},
        )
        assert resp.status_code == 200
        assert "Náhled importu".encode() in resp.data

    def test_accepts_v2_dict_with_users(self, app, admin_client):
        _make_master_event(app)
        csrf = _get_csrf(admin_client, "/import/events/")
        payload = {
            "version": 2,
            "users": [
                {
                    "gs_name": "Gajda Adam",
                    "name": "Adam Gajda",
                    "email": "adam@test.com",
                    "phone": "123",
                    "is_zdravotnik": False,
                }
            ],
            "events": [_minimal_event()],
        }
        resp = admin_client.post(
            "/import/events/preview",
            data={"json_data": json.dumps(payload), "csrf_token": csrf},
        )
        assert resp.status_code == 200
        assert b"Adam Gajda" in resp.data
        assert "Nový".encode() in resp.data

    def test_marks_existing_user_by_name(self, app, admin_client):
        _make_user(app, "Adam Gajda", "adam_existing@test.com")
        csrf = _get_csrf(admin_client, "/import/events/")
        payload = {
            "version": 2,
            "users": [
                {
                    "gs_name": "Gajda Adam",
                    "name": "Adam Gajda",
                    "email": "adam_new@test.com",
                    "phone": None,
                    "is_zdravotnik": False,
                }
            ],
            "events": [],
        }
        resp = admin_client.post(
            "/import/events/preview",
            data={"json_data": json.dumps(payload), "csrf_token": csrf},
        )
        assert resp.status_code == 200
        assert b"Existuje" in resp.data

    def test_marks_existing_user_by_email(self, app, admin_client):
        _make_user(app, "Different Name", "adam@test.com")
        csrf = _get_csrf(admin_client, "/import/events/")
        payload = {
            "version": 2,
            "users": [
                {
                    "gs_name": "Gajda Adam",
                    "name": "Adam Gajda",
                    "email": "adam@test.com",
                    "phone": None,
                    "is_zdravotnik": False,
                }
            ],
            "events": [],
        }
        resp = admin_client.post(
            "/import/events/preview",
            data={"json_data": json.dumps(payload), "csrf_token": csrf},
        )
        assert resp.status_code == 200
        assert b"Existuje" in resp.data

    def test_invalid_json_shows_error(self, admin_client):
        csrf = _get_csrf(admin_client, "/import/events/")
        resp = admin_client.post(
            "/import/events/preview",
            data={"json_data": "not json", "csrf_token": csrf},
        )
        assert resp.status_code == 200
        assert b"Neplatn" in resp.data


# ── Confirm: user creation tests ──────────────────────────────────────────────


class TestImportConfirmUsers:
    def test_creates_user_with_member_role(self, app, admin_client):
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "gs_name": "Gajda Adam",
                    "name": "Adam Gajda",
                    "email": "adam_new@test.com",
                    "phone": "123",
                    "is_zdravotnik": False,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            user = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "adam_new@test.com"))
            assert user is not None
            assert user.is_active is True
            assert any(r.name == "Member" for r in user.roles)

    def test_creates_zdravotnik_user_with_correct_qual(self, app, admin_client):
        me_id = _make_master_event(app)
        with app.app_context():
            q = Qualification(name="Zdravotník")
            db.session.add(q)
            db.session.commit()

        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "gs_name": "Novák Jan",
                    "name": "Jan Novák",
                    "email": "jan@test.com",
                    "phone": "",
                    "is_zdravotnik": True,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            user = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "jan@test.com"))
            assert user is not None
            assert any("zdravotník" in q.name.lower() for q in user.qualifications)

    def test_skips_user_matching_by_name(self, app, admin_client):
        _make_user(app, "Adam Gajda", "adam_orig@test.com")
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "gs_name": "Gajda Adam",
                    "name": "Adam Gajda",
                    "email": "adam_new@test.com",
                    "phone": "",
                    "is_zdravotnik": False,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            count = db.session.scalar(db.select(db.func.count()).where(UserAccount.name == "Adam Gajda"))
            assert count == 1  # not duplicated

    def test_skips_user_matching_by_email(self, app, admin_client):
        _make_user(app, "Different Name", "adam@test.com")
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "gs_name": "Gajda Adam",
                    "name": "Adam Gajda",
                    "email": "adam@test.com",
                    "phone": "",
                    "is_zdravotnik": False,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            count = db.session.scalar(db.select(db.func.count()).where(UserAccount.email == "adam@test.com"))
            assert count == 1

    def test_skips_user_without_email(self, app, admin_client):
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[{"name": "No Email User", "email": "", "phone": "", "is_zdravotnik": False}],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            count = db.session.scalar(db.select(db.func.count()).where(UserAccount.name == "No Email User"))
            assert count == 0

    def test_user_not_created_when_include_unchecked(self, app, admin_client):
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "name": "Unchecked User",
                    "email": "unchecked@test.com",
                    "phone": "",
                    "is_zdravotnik": False,
                    "include": False,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            count = db.session.scalar(db.select(db.func.count()).where(UserAccount.email == "unchecked@test.com"))
            assert count == 0


# ── Confirm: spots and assignments tests ──────────────────────────────────────


class TestImportConfirmSpots:
    def test_standard_3_spots_no_signups(self, app, admin_client):
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Test akce"))
            assert event is not None
            assert len(event.spots) == 3
            mandatory = [s for s in event.spots if not s.is_optional]
            optional = [s for s in event.spots if s.is_optional]
            assert len(mandatory) == 2
            assert len(optional) == 1

    def test_standard_3_spots_with_1_signup(self, app, admin_client):
        """1 signup ≤ 3 → still standard 3-spot pattern."""
        me_id = _make_master_event(app)
        _make_user(app, "Adam Gajda", "adam@test.com")
        ev = _minimal_event()
        ev["signups"] = ["Adam Gajda"]
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Test akce"))
            assert event is not None
            assert len(event.spots) == 3

    def test_dynamic_spots_for_4_signups(self, app, admin_client):
        """4 signups → 1 Zdravotník + 4 Zelenáč = 5 spots."""
        me_id = _make_master_event(app)
        signup_names = [f"User{k} Test" for k in range(4)]
        for k in range(4):
            _make_user(app, f"User{k} Test", f"user{k}@test.com")

        ev = _minimal_event(name="Big Event")
        ev["signups"] = signup_names
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Big Event"))
            assert event is not None
            assert len(event.spots) == 5
            assert all(not s.is_optional for s in event.spots)

    def test_no_spots_when_time_missing(self, app, admin_client):
        me_id = _make_master_event(app)
        ev = _minimal_event(name="No Time Event")
        ev["time_missing"] = True
        ev["start_time"] = None
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "No Time Event"))
            assert event is not None
            assert len(event.spots) == 0


class TestImportCancelledEvents:
    def test_cancelled_event_gets_cancelled_status(self, app, admin_client):

        me_id = _make_master_event(app)
        ev = _minimal_event(name="Zrušená akce", date="2030-06-01")
        ev["cancelled"] = True
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Zrušená akce"))
            assert event is not None
            assert event.status == EventStatus.CANCELLED

    def test_cancelled_event_has_no_spots(self, app, admin_client):
        me_id = _make_master_event(app)
        ev = _minimal_event(name="Zrušená bez pozic")
        ev["cancelled"] = True
        ev["signups"] = ["Novák Jan"]
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Zrušená bez pozic"))
            assert event is not None
            assert len(event.spots) == 0

    def test_past_cancelled_event_stays_cancelled(self, app, admin_client):
        """A past event that is cancelled should not be overridden to COMPLETED."""

        me_id = _make_master_event(app)
        ev = _minimal_event(name="Stará zrušená akce", date="2020-01-01")
        ev["cancelled"] = True
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Stará zrušená akce"))
            assert event is not None
            assert event.status == EventStatus.CANCELLED

    def test_non_cancelled_event_unaffected(self, app, admin_client):

        me_id = _make_master_event(app)
        ev = _minimal_event(name="Normální akce")
        ev["cancelled"] = False
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Normální akce"))
            assert event is not None
            assert event.status == EventStatus.DRAFT


class TestImportConfirmAssignments:
    def test_rp_assigned_to_zdravotnik_spot(self, app, admin_client):
        me_id = _make_master_event(app)
        rp_id = _make_user(app, "Roman Vykydal", "rp@test.com")

        ev = _minimal_event(name="RP Event")
        ev["responsible_person_id"] = rp_id
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "RP Event"))
            assert event is not None
            zdravotnik_spot = next(s for s in event.spots if s.description == "Zdravotník")
            assert zdravotnik_spot.assignment is not None
            rp_user = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "rp@test.com"))
            assert zdravotnik_spot.assignment.user_id == rp_user.id

    def test_signups_assigned_to_zelenac_spots(self, app, admin_client):
        me_id = _make_master_event(app)
        _make_user(app, "Adam Gajda", "adam@test.com")
        _make_user(app, "Marek Skyba", "marek@test.com")

        ev = _minimal_event(name="Signup Event")
        ev["signups"] = ["Adam Gajda", "Marek Skyba"]
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Signup Event"))
            assert event is not None
            assignments = [s.assignment for s in event.spots if s.description == "Zelenáč" and s.assignment is not None]
            assert len(assignments) == 2

    def test_signup_without_user_account_is_skipped(self, app, admin_client):
        """A signup name that doesn't match any user in DB simply doesn't get assigned."""
        me_id = _make_master_event(app)
        ev = _minimal_event(name="Unknown Signup")
        ev["signups"] = ["Neexistující Uživatel"]
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Unknown Signup"))
            assert event is not None
            # Spots are created but none are assigned
            assert all(s.assignment is None for s in event.spots)

    def test_newly_imported_user_gets_assigned(self, app, admin_client):
        """A user created in the same import can be assigned to an event signup."""
        me_id = _make_master_event(app)
        ev = _minimal_event(name="New User Event")
        ev["signups"] = ["Adam Gajda"]
        resp = _post_confirm(
            app,
            admin_client,
            events=[ev],
            users=[
                {
                    "gs_name": "Gajda Adam",
                    "name": "Adam Gajda",
                    "email": "adam@test.com",
                    "phone": "",
                    "is_zdravotnik": False,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "New User Event"))
            assert event is not None
            assigned_names = [s.assignment.user.name for s in event.spots if s.assignment is not None]
            assert "Adam Gajda" in assigned_names

    def test_past_event_gets_auto_debriefing(self, app, admin_client):
        """Past events (end < now) should automatically receive DebriefingRecord rows."""

        me_id = _make_master_event(app)
        uid = _make_user(app, "Eva Nováková", "eva@test.com")
        ev = _minimal_event(name="Past Event Debrief", date="2020-01-15")
        ev["responsible_person_id"] = uid
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Past Event Debrief"))
            assert event is not None
            rp_spot = next((s for s in event.spots if s.assignment is not None), None)
            assert rp_spot is not None
            asgn_id = rp_spot.assignment.id
            debrief = db.session.scalar(db.select(DebriefingRecord).where(DebriefingRecord.assignment_id == asgn_id))
            assert debrief is not None
            assert "importovaný" in debrief.feedback_event.lower()

    def test_duplicate_event_skipped_when_unchecked(self, app, admin_client):
        """If admin unchecks a duplicate event row, it must not be created."""
        me_id = _make_master_event(app)
        ev1 = _minimal_event(name="Dup Event")
        resp = _post_confirm(app, admin_client, events=[ev1], master_event_id=me_id)
        assert resp.status_code == 302

        # Re-submit the same event but with include=0 (unchecked by admin)
        csrf = _get_csrf(admin_client, "/import/events/")
        data = {
            "csrf_token": csrf,
            "event_count": "1",
            "user_count": "0",
            "global_master_event_id": str(me_id),
            "ev_0_include": "",  # not checked
            "ev_0_name": "Dup Event",
            "ev_0_date": "2030-05-01",
            "ev_0_start_time": "10:00",
            "ev_0_end_time": "12:00",
            "ev_0_location": "",
            "ev_0_paid": "",
            "ev_0_contact_person": "",
            "ev_0_description": "",
            "ev_0_time_missing": "0",
            "ev_0_responsible_person_id": "",
            "ev_0_signup_count": "0",
        }
        resp2 = admin_client.post("/import/events/confirm", data=data, follow_redirects=False)
        assert resp2.status_code == 302
        with app.app_context():
            count = db.session.scalar(db.select(db.func.count()).select_from(Event).where(Event.name == "Dup Event"))
            assert count == 1  # still only 1, not 2


class TestImportIdempotency:
    def test_rerun_does_not_duplicate_user(self, app, admin_client):
        """Importing the same user payload twice creates the user only once."""
        me_id = _make_master_event(app)
        user_payload = [
            {
                "gs_name": "Novák Jan",
                "name": "Novák Jan",
                "email": "jan@test.com",
                "phone": None,
                "is_zdravotnik": False,
            }
        ]

        for _ in range(2):
            _post_confirm(
                app, admin_client, events=[_minimal_event(name=f"Akce {_}")], users=user_payload, master_event_id=me_id
            )

        with app.app_context():
            count = db.session.scalar(
                db.select(db.func.count())
                .select_from(UserAccount)
                .where(db.func.lower(UserAccount.name) == "novák jan")
            )
            assert count == 1

    def test_rerun_does_not_duplicate_user_by_email(self, app, admin_client):
        """If a user already exists with the same email, it must not be re-created."""
        me_id = _make_master_event(app)
        _make_user(app, "Petra Horáková", "petra@test.com")
        user_payload = [
            {
                "gs_name": "Horáková Petra",
                "name": "Petra Horáková NEW",
                "email": "petra@test.com",
                "phone": None,
                "is_zdravotnik": False,
            }
        ]

        _post_confirm(app, admin_client, events=[_minimal_event()], users=user_payload, master_event_id=me_id)

        with app.app_context():
            count = db.session.scalar(
                db.select(db.func.count()).select_from(UserAccount).where(UserAccount.email == "petra@test.com")
            )
            assert count == 1  # not duplicated despite different name in payload


# ── is_ridic (Řidič sanitky) import tests ─────────────────────────────────────


@pytest.fixture()
def ensure_ridic_qual(app):
    """Ensure the 'Řidič sanitky' qualification exists in the test DB.

    Function-scoped because clean_db truncates the qualification table after each test.
    """
    with app.app_context():
        exists = db.session.scalar(db.select(Qualification).where(Qualification.name == "Řidič sanitky"))
        if not exists:
            db.session.add(Qualification(name="Řidič sanitky"))
            db.session.commit()


@pytest.fixture()
def ensure_zelenac_qual(app):
    """Ensure the 'Zelenáč' qualification exists in the test DB.

    Function-scoped because clean_db truncates the qualification table after each test.
    """
    with app.app_context():
        exists = db.session.scalar(db.select(Qualification).where(Qualification.name == "Zelenáč"))
        if not exists:
            db.session.add(Qualification(name="Zelenáč"))
            db.session.commit()


class TestImportScriptIsRidic:
    """Unit tests for is_ridic extraction in scripts/import_events.py."""

    def test_extract_users_includes_is_ridic(self):
        """extract_users() returns is_ridic from Lidi column F."""

        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        users = _script.extract_users(wb)
        by_name = {u["name"]: u for u in users}
        # Kratochvíl Tomáš: zdravotník=True, ridic=True
        assert by_name["Kratochvíl Tomáš"]["is_zdravotnik"] is True
        assert by_name["Kratochvíl Tomáš"]["is_ridic"] is True
        # Svoboda Petr: zdravotník=False, ridic=False
        assert by_name["Svoboda Petr"]["is_zdravotnik"] is False
        assert by_name["Svoboda Petr"]["is_ridic"] is False
        # Horáková Marie: zdravotník=False, ridic=True
        assert by_name["Horáková Marie"]["is_zdravotnik"] is False
        assert by_name["Horáková Marie"]["is_ridic"] is True

    def test_extract_users_ridic_false_by_default_when_not_in_lidi(self):
        """Person not in Lidi gets is_ridic=False."""

        fixture = Path(__file__).parent / "fixtures" / "test_import.xlsx"
        wb = openpyxl.load_workbook(str(fixture), data_only=True)
        users = _script.extract_users(wb)
        by_name = {u["name"]: u for u in users}
        # Pokorný Zdeněk is only in Dozory, not in Lidi
        assert by_name["Pokorný Zdeněk"]["is_ridic"] is False


class TestImportConfirmRidic:
    """Integration tests for is_ridic support in the import confirm route."""

    def test_creates_ridic_only_user_with_ridic_and_zelenac(
        self, app, admin_client, ensure_ridic_qual, ensure_zelenac_qual
    ):
        """User with is_ridic=True, is_zdravotnik=False gets Řidič sanitky + Zelenáč."""
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "gs_name": "Horáková Marie",
                    "name": "Marie Horáková",
                    "email": "marie@test.com",
                    "phone": "",
                    "is_zdravotnik": False,
                    "is_ridic": True,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            user = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "marie@test.com"))
            assert user is not None
            qual_names = {q.name for q in user.qualifications}
            assert "Řidič sanitky" in qual_names
            assert "Zelenáč" in qual_names
            assert "Zdravotník" not in qual_names

    def test_creates_user_with_both_quals(self, app, admin_client, ensure_ridic_qual):
        """User with is_zdravotnik=True, is_ridic=True gets both Zdravotník and Řidič sanitky."""
        me_id = _make_master_event(app)
        with app.app_context():
            if not db.session.scalar(db.select(Qualification).where(Qualification.name == "Zdravotník")):
                db.session.add(Qualification(name="Zdravotník"))
                db.session.commit()
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "gs_name": "Kratochvíl Tomáš",
                    "name": "Tomáš Kratochvíl",
                    "email": "tomas@test.com",
                    "phone": "",
                    "is_zdravotnik": True,
                    "is_ridic": True,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            user = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "tomas@test.com"))
            assert user is not None
            qual_names = {q.name for q in user.qualifications}
            assert "Zdravotník" in qual_names
            assert "Řidič sanitky" in qual_names

    def test_creates_neither_gets_zelenac(self, app, admin_client, ensure_zelenac_qual):
        """User with is_zdravotnik=False, is_ridic=False gets Zelenáč."""
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "gs_name": "Svoboda Petr",
                    "name": "Petr Svoboda",
                    "email": "petr@test.com",
                    "phone": "",
                    "is_zdravotnik": False,
                    "is_ridic": False,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            user = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "petr@test.com"))
            assert user is not None
            qual_names = {q.name for q in user.qualifications}
            assert "Zelenáč" in qual_names
            assert "Řidič sanitky" not in qual_names

    def test_updates_existing_user_adds_ridic_qual(self, app, admin_client, ensure_ridic_qual, ensure_zelenac_qual):
        """Existing user without Řidič sanitky gets it added when import says is_ridic=True."""
        # Create user with only Zelenáč
        with app.app_context():
            zelenac = db.session.scalar(db.select(Qualification).where(Qualification.name == "Zelenáč"))
            u = _conftest_make_user("existing_ridic@test.com", "Jan Řidič", Role.MEMBER)
            if zelenac:
                u.qualifications = [zelenac]
            db.session.commit()
            user_id = str(u.id)
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "db_id": user_id,
                    "gs_name": "Řidič Jan",
                    "name": "Jan Řidič",
                    "email": "existing_ridic@test.com",
                    "phone": "",
                    "is_zdravotnik": False,
                    "is_ridic": True,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            user = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "existing_ridic@test.com"))
            qual_names = {q.name for q in user.qualifications}
            assert "Řidič sanitky" in qual_names
            assert "Zelenáč" in qual_names

    def test_updates_existing_user_removes_zelenac_adds_zdravotnik(self, app, admin_client, ensure_zelenac_qual):
        """Existing Zelenáč gets Zdravotník when import says is_zdravotnik=True."""
        with app.app_context():
            zelenac = db.session.scalar(db.select(Qualification).where(Qualification.name == "Zelenáč"))
            if not db.session.scalar(db.select(Qualification).where(Qualification.name == "Zdravotník")):
                db.session.add(Qualification(name="Zdravotník"))
                db.session.commit()
            u = _conftest_make_user("was_zelenac@test.com", "Jana Zelenáčová", Role.MEMBER)
            if zelenac:
                u.qualifications = [zelenac]
            db.session.commit()
            user_id = str(u.id)
        me_id = _make_master_event(app)
        resp = _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "db_id": user_id,
                    "gs_name": "Zelenáčová Jana",
                    "name": "Jana Zelenáčová",
                    "email": "was_zelenac@test.com",
                    "phone": "",
                    "is_zdravotnik": True,
                    "is_ridic": False,
                }
            ],
            master_event_id=me_id,
        )
        assert resp.status_code == 302
        with app.app_context():
            user = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "was_zelenac@test.com"))
            qual_names = {q.name for q in user.qualifications}
            assert "Zdravotník" in qual_names
            assert "Zelenáč" not in qual_names

    def test_no_update_when_quals_already_match(self, app, admin_client, ensure_ridic_qual):
        """Existing user with correct quals is not modified (no audit entry added)."""
        with app.app_context():
            ridic = db.session.scalar(db.select(Qualification).where(Qualification.name == "Řidič sanitky"))
            u = _conftest_make_user("already_ridic@test.com", "Already Driver", Role.MEMBER)
            if ridic:
                u.qualifications = [ridic]
            db.session.commit()
            user_id = str(u.id)
            initial_audit_count = db.session.scalar(
                db.select(db.func.count()).select_from(AuditLogEntry).where(AuditLogEntry.entity_id == str(u.id))
            )
        me_id = _make_master_event(app)
        _post_confirm(
            app,
            admin_client,
            events=[_minimal_event()],
            users=[
                {
                    "db_id": user_id,
                    "gs_name": "Driver Already",
                    "name": "Already Driver",
                    "email": "already_ridic@test.com",
                    "phone": "",
                    "is_zdravotnik": False,
                    "is_ridic": True,
                }
            ],
            master_event_id=me_id,
        )
        with app.app_context():
            final_audit_count = db.session.scalar(
                db.select(db.func.count()).select_from(AuditLogEntry).where(AuditLogEntry.entity_id == user_id)
            )
            assert final_audit_count == initial_audit_count  # no new audit entry


class TestImportPreviewRidic:
    """Tests for is_ridic display in the preview route."""

    def test_preview_shows_ridic_badge_for_ridic_user(self, app, admin_client):
        """Preview shows 'Řidič sanitky' badge when is_ridic=True."""
        csrf = _get_csrf(admin_client, "/import/events/")
        payload = {
            "version": 2,
            "users": [
                {
                    "gs_name": "Horáková Marie",
                    "name": "Marie Horáková",
                    "email": "marie@test.com",
                    "phone": None,
                    "is_zdravotnik": False,
                    "is_ridic": True,
                }
            ],
            "events": [],
        }
        resp = admin_client.post(
            "/import/events/preview",
            data={"json_data": json.dumps(payload), "csrf_token": csrf},
        )
        assert resp.status_code == 200
        assert "Řidič sanitky".encode() in resp.data

    def test_preview_shows_qual_update_badge_for_existing_user(
        self, app, admin_client, ensure_zelenac_qual, ensure_ridic_qual
    ):
        """Preview shows 'Aktualizace kvalifikací' when existing user's quals differ."""
        with app.app_context():
            zelenac = db.session.scalar(db.select(Qualification).where(Qualification.name == "Zelenáč"))
            u = _conftest_make_user("preview_qual@test.com", "Preview User", Role.MEMBER)
            if zelenac:
                u.qualifications = [zelenac]
            db.session.commit()
        csrf = _get_csrf(admin_client, "/import/events/")
        payload = {
            "version": 2,
            "users": [
                {
                    "gs_name": "User Preview",
                    "name": "Preview User",
                    "email": "preview_qual@test.com",
                    "phone": None,
                    "is_zdravotnik": False,
                    "is_ridic": True,
                }
            ],
            "events": [],
        }
        resp = admin_client.post(
            "/import/events/preview",
            data={"json_data": json.dumps(payload), "csrf_token": csrf},
        )
        assert resp.status_code == 200
        assert "Aktualizace kvalifikací".encode() in resp.data


class TestImportAutoClose:
    def test_fully_occupied_future_event_gets_assignments_closed(self, app, admin_client):
        me_id = _make_master_event(app)
        rp_id = _make_user(app, "Roman Vykydal", "rp@test.com")
        _make_user(app, "Adam Gajda", "adam@test.com")

        ev = _minimal_event(name="Plně obsazená akce")
        ev["responsible_person_id"] = rp_id
        ev["signups"] = ["Adam Gajda"]
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Plně obsazená akce"))
            assert event is not None
            assert event.status == EventStatus.ASSIGNMENTS_CLOSED

    def test_partially_filled_future_event_stays_draft(self, app, admin_client):
        me_id = _make_master_event(app)
        rp_id = _make_user(app, "Roman Vykydal", "rp2@test.com")

        ev = _minimal_event(name="Částečně obsazená akce")
        ev["responsible_person_id"] = rp_id
        ev["signups"] = []
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Částečně obsazená akce"))
            assert event is not None
            assert event.status == EventStatus.DRAFT

    def test_no_assignments_future_event_stays_draft(self, app, admin_client):
        me_id = _make_master_event(app)
        ev = _minimal_event(name="Prázdná akce")
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Prázdná akce"))
            assert event is not None
            assert event.status == EventStatus.DRAFT

    def test_fully_occupied_past_event_stays_completed(self, app, admin_client):
        me_id = _make_master_event(app)
        rp_id = _make_user(app, "Eva Nováková", "eva2@test.com")
        _make_user(app, "Petr Novák", "petr2@test.com")

        ev = _minimal_event(name="Historická plná akce", date="2020-01-15")
        ev["responsible_person_id"] = rp_id
        ev["signups"] = ["Petr Novák"]
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Historická plná akce"))
            assert event is not None
            assert event.status == EventStatus.COMPLETED

    def test_fully_occupied_dynamic_spots_future_event_gets_assignments_closed(self, app, admin_client):
        me_id = _make_master_event(app)
        rp_id = _make_user(app, "Roman Vykydal", "rp3@test.com")
        signup_names = [f"User{k} Big" for k in range(4)]
        for k in range(4):
            _make_user(app, f"User{k} Big", f"user_big{k}@test.com")

        ev = _minimal_event(name="Velká plná akce")
        ev["responsible_person_id"] = rp_id
        ev["signups"] = signup_names
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Velká plná akce"))
            assert event is not None
            assert event.status == EventStatus.ASSIGNMENTS_CLOSED

    def test_unknown_signup_leaves_event_draft(self, app, admin_client):
        me_id = _make_master_event(app)
        rp_id = _make_user(app, "Roman Vykydal", "rp4@test.com")

        ev = _minimal_event(name="Neznámý přihlášený")
        ev["responsible_person_id"] = rp_id
        ev["signups"] = ["Neexistující Osoba"]
        resp = _post_confirm(app, admin_client, events=[ev], master_event_id=me_id)
        assert resp.status_code == 302
        with app.app_context():
            event = db.session.scalar(db.select(Event).where(Event.name == "Neznámý přihlášený"))
            assert event is not None
            assert event.status == EventStatus.DRAFT
