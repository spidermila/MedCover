"""Tests for the výkaz práce (employee work report) feature."""

import os
from datetime import datetime, timedelta, timezone

import openpyxl
import pytest

from app.extensions import db
from app.models.assignment import Assignment
from app.models.event import Event, EventSpot, EventStatus
from app.models.master_event import MasterEvent
from app.models.role import Role
from app.models.user import UserAccount
from app.scheduler_tasks import cleanup_work_report_files
from app.work_report_generator import generate_work_report
from tests.conftest import _login, _make_user


def _make_paid_event(
    user: UserAccount,
    start: datetime,
    end: datetime,
    actual_hours: float | None = None,
    name: str = "Testovací akce",
) -> Event:
    """Create a COMPLETED paid event with an assignment for *user*."""
    me = MasterEvent(name="Testovací ME")
    db.session.add(me)
    db.session.flush()

    ev = Event(
        name=name,
        master_event_id=me.id,
        status=EventStatus.COMPLETED,
        start_datetime=start,
        end_datetime=end,
        paid=True,
    )
    if actual_hours is not None:
        ev.actual_start_datetime = start
        ev.actual_end_datetime = start + timedelta(hours=actual_hours)
    db.session.add(ev)
    db.session.flush()

    spot = EventSpot(event_id=ev.id, description="Záchranář")
    db.session.add(spot)
    db.session.flush()

    admin = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "admin@test.com"))
    if admin is None:
        role = db.session.scalar(db.select(Role).where(Role.name == Role.ADMIN))
        admin = UserAccount(email="admin@test.com", name="Admin", is_active=True)
        admin.set_password("adminpass")
        admin.roles = [role]
        db.session.add(admin)
        db.session.flush()

    asgn = Assignment(spot_id=spot.id, user_id=user.id, assigned_by_id=admin.id)
    db.session.add(asgn)
    db.session.commit()
    return ev


# ── Route smoke tests ──────────────────────────────────────────────────────────


class TestVykazIndex:
    def test_requires_login(self, client):
        resp = client.get("/work-report/", follow_redirects=False)
        assert resp.status_code in (301, 302)
        assert "/auth/login" in resp.headers["Location"]

    def test_index_returns_200(self, app, client):
        with app.app_context():
            _make_user("vykaz_idx@test.com", "Vykaz User", Role.MEMBER)
        _login(client, "vykaz_idx@test.com")
        resp = client.get("/work-report/")
        assert resp.status_code == 200
        assert "Výkaz práce" in resp.data.decode()
        assert "Leden" in resp.data.decode()
        assert "Prosinec" in resp.data.decode()

    def test_viewer_gets_403(self, app, client):
        with app.app_context():
            _make_user("vykaz_viewer@test.com", "Viewer", Role.VIEWER)
        _login(client, "vykaz_viewer@test.com")
        resp = client.get("/work-report/", follow_redirects=False)
        assert resp.status_code == 403


class TestVykazGenerate:
    def test_invalid_month_rejected(self, app, client):
        with app.app_context():
            _make_user("vykaz_bad@test.com", "Vykaz User", Role.MEMBER)
        _login(client, "vykaz_bad@test.com")
        resp = client.post(
            "/work-report/generate",
            data={"year": "2026", "month": "99", "csrf_token": "x"},
            follow_redirects=True,
        )
        assert resp.status_code == 200
        assert "Měsíc musí být" in resp.data.decode()

    def test_generate_creates_file_and_shows_list(self, app, client, tmp_path, monkeypatch):
        """POST /work-report/generate creates an xlsx and shows it in the report list."""
        with app.app_context():
            _make_user("vykaz_gen@test.com", "Jana Nováková", Role.MEMBER)

        # Patch instance_path so we don't litter the real instance dir
        monkeypatch.setattr(app, "instance_path", str(tmp_path))

        _login(client, "vykaz_gen@test.com")
        resp = client.post(
            "/work-report/generate",
            data={"year": "2026", "month": "1", "csrf_token": "x"},
            follow_redirects=True,
        )
        assert resp.status_code == 200
        body = resp.data.decode()
        # Flash success message
        assert "Leden" in body
        assert "2026" in body
        # Index now shows the download link for the generated report
        assert "/work-report/download" in body

        with app.app_context():
            user = db.session.scalar(db.select(UserAccount).where(UserAccount.email == "vykaz_gen@test.com"))
            out = tmp_path / "work_report" / str(user.id) / "2026-01.xlsx"
        assert out.exists(), "xlsx file was not created"


class TestVykazDownload:
    def test_download_missing_file_redirects(self, app, client):
        with app.app_context():
            _make_user("vykaz_dl@test.com", "Vykaz User", Role.MEMBER)
        _login(client, "vykaz_dl@test.com")
        resp = client.get("/work-report/download?year=2026&month=1", follow_redirects=True)
        assert resp.status_code == 200
        assert "nenalezen" in resp.data.decode()


# ── Generator unit tests ───────────────────────────────────────────────────────


class TestVykazGenerator:
    def test_generator_produces_valid_xlsx(self, app, tmp_path, monkeypatch):
        """generate_work_report creates a readable xlsx with correct sheet name."""

        with app.app_context():
            monkeypatch.setattr(app, "instance_path", str(tmp_path))
            u = _make_user("vykaz_unit@test.com", "Petr Svoboda", Role.MEMBER)
            path = generate_work_report(u, 2026, 1)

        assert path.exists()
        wb = openpyxl.load_workbook(str(path))
        assert wb.sheetnames == ["Leden"]

    def test_generator_correct_day_count(self, app, tmp_path, monkeypatch):
        """February 2026 should have 28 day rows (not 29 or 31)."""

        with app.app_context():
            monkeypatch.setattr(app, "instance_path", str(tmp_path))
            u = _make_user("vykaz_feb@test.com", "Vykaz User", Role.MEMBER)
            path = generate_work_report(u, 2026, 2)

        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # Row 10 = day 1, row 10+27 = day 28; row 38 should be total row
        assert ws.cell(row=10, column=1).value == 1
        assert ws.cell(row=37, column=1).value == 28
        assert ws.cell(row=38, column=1).value == "Celkem hodin"

    def test_generator_fills_paid_events(self, app, tmp_path, monkeypatch):
        """Events attended by the user appear in the correct day row."""

        now = datetime(2026, 3, 15, 10, 0, tzinfo=timezone.utc)
        end = datetime(2026, 3, 15, 14, 0, tzinfo=timezone.utc)

        with app.app_context():
            monkeypatch.setattr(app, "instance_path", str(tmp_path))
            u = _make_user("vykaz_ev@test.com", "Vykaz User", Role.MEMBER)
            _make_paid_event(u, now, end, actual_hours=4.0, name="Hasiči 2026")
            path = generate_work_report(u, 2026, 3)

        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # March 15 is day 15 → row 10 + 14 = 24
        day_row = 10 + 15 - 1
        assert ws.cell(row=day_row, column=3).value == pytest.approx(4.0)
        assert "Hasiči 2026" in (ws.cell(row=day_row, column=4).value or "")

    def test_generator_holiday_yellow(self, app, tmp_path, monkeypatch):
        """January 1 (Czech public holiday) should have yellow fill."""

        with app.app_context():
            monkeypatch.setattr(app, "instance_path", str(tmp_path))
            u = _make_user("vykaz_hol@test.com", "Vykaz User", Role.MEMBER)
            path = generate_work_report(u, 2026, 1)

        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        cell_a10 = ws.cell(row=10, column=1)
        assert cell_a10.fill.fgColor.rgb == "FFFFFF00", "Jan 1 must have yellow fill"

    def test_generator_weekend_red_font(self, app, tmp_path, monkeypatch):
        """Saturday day-name cell should use red font."""

        with app.app_context():
            monkeypatch.setattr(app, "instance_path", str(tmp_path))
            u = _make_user("vykaz_wknd@test.com", "Vykaz User", Role.MEMBER)
            # January 2026: day 3 = Saturday
            path = generate_work_report(u, 2026, 1)

        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # day 3 = row 12, col B
        b12 = ws.cell(row=12, column=2)
        assert b12.value == "SO"
        assert b12.font.color.rgb == "FFFF0000", "Saturday must have red font"


class TestCleanupVykazFiles:
    def test_cleanup_removes_old_files(self, tmp_path):

        work_report_dir = tmp_path / "work_report" / "user1"
        work_report_dir.mkdir(parents=True)
        old_file = work_report_dir / "2025-01.xlsx"
        old_file.write_bytes(b"x")
        # backdate mtime to 2 days ago
        two_days_ago = datetime.now(timezone.utc) - timedelta(days=2)
        os.utime(old_file, (two_days_ago.timestamp(), two_days_ago.timestamp()))

        new_file = work_report_dir / "2026-01.xlsx"
        new_file.write_bytes(b"x")  # fresh mtime

        removed = cleanup_work_report_files(str(tmp_path))
        assert removed == 1
        assert not old_file.exists()
        assert new_file.exists()
