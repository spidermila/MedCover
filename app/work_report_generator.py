"""
Výkaz práce (employee work report) xlsx generator.

Generates a single-sheet openpyxl workbook that matches the layout of the
legacy Google-Sheets "Dozory YYYY.xlsx" monthly report used by Czech Red
Cross members to document worked hours for DPP payroll purposes.

Public entry point::

    path = generate_work_report(user, year, month)

The file is written to  instance/work_report/<user_id>/<year>-<MM>.xlsx
and is overwritten on each call.  Callers are responsible for serving
the file and scheduling cleanup (files older than 1 day should be removed).
"""

import calendar
import io
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING

import holidays
import sqlalchemy as sa
from flask import current_app
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

if TYPE_CHECKING:
    from app.models import UserAccount

# ── Czech locale constants ────────────────────────────────────────────────────

CZ_MONTH_NAMES = [
    "",
    "Leden",
    "Únor",
    "Březen",
    "Duben",
    "Květen",
    "Červen",
    "Červenec",
    "Srpen",
    "Září",
    "Říjen",
    "Listopad",
    "Prosinec",
]

CZ_WEEKDAY_ABBR = ["PO", "ÚT", "ST", "ČT", "PÁ", "SO", "NE"]  # Mon=0 … Sun=6

# ── Colours ──────────────────────────────────────────────────────────────────

_BLUE_FILL = PatternFill("solid", fgColor="FF99CCFF")  # column header row
_CYAN_FILL = PatternFill("solid", fgColor="FFCCFFFF")  # info block labels
_YELLOW_FILL = PatternFill("solid", fgColor="FFFFFF00")  # public holiday
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFFFF")  # normal day
_RED_FONT = Font(name="Calibri", size=10, color="FFFF0000")  # weekend day name
_STD_FONT = Font(name="Calibri", size=10)
_BOLD_FONT = Font(name="Calibri", size=10, bold=True)

# ── Borders ──────────────────────────────────────────────────────────────────


def _side(style: str) -> Side:
    return Side(border_style=style)


_HDR_BORDER_A = Border(
    left=_side("medium"),
    right=_side("thin"),
    top=_side("medium"),
    bottom=_side("medium"),
)
_HDR_BORDER_B = Border(
    left=_side("medium"),
    right=_side("thin"),
    top=_side("medium"),
)
_HDR_BORDER_C = Border(
    left=_side("medium"),
    right=_side("thin"),
    top=_side("medium"),
    bottom=_side("medium"),
)
_HDR_BORDER_D = Border(
    left=_side("thin"),
    right=_side("medium"),
    top=_side("medium"),
    bottom=_side("medium"),
)
# Right-edge cell (E) of merged D:E — only right/top/bottom needed
_HDR_BORDER_E = Border(
    right=_side("medium"),
    top=_side("medium"),
    bottom=_side("medium"),
)
_DAY_BORDER_A = Border(
    left=_side("medium"),
    right=_side("thin"),
    top=_side("thin"),
    bottom=_side("thin"),
)
_DAY_BORDER_B = Border(
    left=_side("thin"),
    right=_side("thin"),
    top=_side("thin"),
    bottom=_side("thin"),
)
_DAY_BORDER_C = Border(
    left=_side("thin"),
    right=_side("thin"),
    top=_side("thin"),
    bottom=_side("thin"),
)
_DAY_BORDER_D = Border(
    left=_side("thin"),
    right=_side("medium"),
    top=_side("thin"),
    bottom=_side("thin"),
)
# Right-edge cell (E) of merged D:E for day rows
_DAY_BORDER_E = Border(
    right=_side("medium"),
    top=_side("thin"),
    bottom=_side("thin"),
)
_TOTAL_BORDER_A = Border(
    left=_side("medium"),
    right=_side("thin"),
    top=_side("medium"),
    bottom=_side("medium"),
)
_TOTAL_BORDER_B = Border(
    left=_side("thin"),
    right=_side("thin"),
    top=_side("medium"),
    bottom=_side("medium"),
)
_TOTAL_BORDER_C = Border(
    left=_side("thin"),
    right=_side("thin"),
    top=_side("medium"),
    bottom=_side("medium"),
)
_TOTAL_BORDER_D = Border(
    left=_side("thin"),
    right=_side("medium"),
    top=_side("medium"),
    bottom=_side("medium"),
)
_TOTAL_BORDER_E = Border(
    right=_side("medium"),
    top=_side("medium"),
    bottom=_side("medium"),
)

# Info block (rows 3-8) borders — matched to sample vykaz.xlsx
# Row 3: title, all-medium box
_INFO_TITLE_A = Border(
    left=_side("medium"),
    right=_side("medium"),
    top=_side("medium"),
    bottom=_side("medium"),
)
_INFO_TITLE_MID = Border(top=_side("medium"), bottom=_side("medium"))
_INFO_TITLE_E = Border(right=_side("medium"), top=_side("medium"), bottom=_side("medium"))
# Rows 4-6: label | spacer-B | spacer-C | value | right-edge
_INFO_LABEL = Border(
    left=_side("medium"),
    right=_side("thin"),
    top=_side("thin"),
    bottom=_side("thin"),
)
_INFO_MID_B = Border(top=_side("thin"), bottom=_side("thin"))
_INFO_MID_C = Border(right=_side("thin"), top=_side("thin"), bottom=_side("thin"))
_INFO_VALUE_D = Border(
    left=_side("thin"),
    right=_side("medium"),
    top=_side("thin"),
    bottom=_side("thin"),
)
_INFO_VALUE_E = Border(right=_side("medium"), top=_side("thin"), bottom=_side("thin"))
# Row 7: empty spacer — D7:E7 merged so no interior border appears between them
_INFO_SPACER_A = Border(left=_side("medium"), top=_side("thin"))
_INFO_SPACER_B = Border(top=_side("thin"))
_INFO_SPACER_C = Border(right=_side("thin"), top=_side("thin"))
_INFO_SPACER_D = Border(
    left=_side("thin"),
    top=_side("thin"),
    bottom=_side("thin"),
)
_INFO_SPACER_E = Border(right=_side("medium"), top=_side("thin"), bottom=_side("thin"))
# Row 8: month / year — bottom=medium acts as separator before column headers
_INFO_MONTH_A = Border(
    left=_side("medium"),
    right=_side("thin"),
    top=_side("thin"),
    bottom=_side("medium"),
)
_INFO_MONTH_B = Border(right=_side("thin"), top=_side("thin"), bottom=_side("medium"))
_INFO_MONTH_C = Border(
    left=_side("thin"),
    right=_side("thin"),
    top=_side("thin"),
    bottom=_side("medium"),
)
_INFO_MONTH_D = Border(
    left=_side("thin"),
    right=_side("thin"),
    top=_side("thin"),
    bottom=_side("medium"),
)
_INFO_MONTH_E = Border(
    left=_side("thin"),
    right=_side("medium"),
    top=_side("thin"),
    bottom=_side("medium"),
)

# ── Layout constants ──────────────────────────────────────────────────────────

_ROW_HEIGHT = 15.75
_COL_WIDTHS = {
    "A": 6.86,  # date number
    "B": 6.43,  # day abbreviation
    "C": 9.43,  # hours
    "D": 20.29,  # description (merged D:E)
    "E": 22.43,
}

_FIRST_DATA_ROW = 10  # row index of day-1 (1-based)
_HEADER_ROW = 9  # column header row

_SIGNATURE_TARGET_HEIGHT_PX = 60
_SIGNATURE_ROW_HEIGHT_PT = 50
# Nudge the image ~0.3 cm to the right of the last _COL_WIDTHS column's right
# edge so it visually clears the column border on the printed page.
_SIGNATURE_RIGHT_NUDGE_CM = 0.3


# ── Helpers ───────────────────────────────────────────────────────────────────


def _apply_row_height(ws: Worksheet, row: int, height: float = _ROW_HEIGHT) -> None:
    ws.row_dimensions[row].height = height


def _write_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: object = None,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def _fetch_events_for_month(user_id: str, year: int, month: int) -> dict[int, tuple[Decimal, list[str]]]:
    """Return {day: (total_hours, [event_names])} for the user's paid completed events."""
    from app.extensions import db  # pylint: disable=import-outside-toplevel
    from app.models import Assignment, Event, EventSpot, EventStatus  # pylint: disable=import-outside-toplevel

    period_start = datetime(year, month, 1, tzinfo=timezone.utc)
    last_day = calendar.monthrange(year, month)[1]
    period_end = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)

    rows = (
        db.session.execute(
            sa.select(Event)
            .join(EventSpot, EventSpot.event_id == Event.id)
            .join(Assignment, Assignment.spot_id == EventSpot.id)
            .where(
                Assignment.user_id == user_id,
                Event.status == EventStatus.COMPLETED,
                Event.paid == sa.true(),
                Event.start_datetime >= period_start,
                Event.start_datetime <= period_end,
            )
        )
        .scalars()
        .all()
    )

    result: dict[int, tuple[Decimal, list[str]]] = {}
    for ev in rows:
        hours = ev.billable_hours
        day = ev.start_datetime.day
        if day in result:
            prev_hours, prev_names = result[day]
            result[day] = (prev_hours + hours, prev_names + [ev.name])
        else:
            result[day] = (hours, [ev.name])
    return result


# ── Core generator ────────────────────────────────────────────────────────────


def _setup_worksheet(ws: Worksheet, month_name: str) -> None:
    """Set column widths, page orientation, margins."""
    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    # A4 portrait, fit-to-page so the report block always fills the page
    # without the user opening Excel's page-setup dialog. Margins match the
    # legacy Google-Sheets "Dozory YYYY.xlsx" template. fitToPage requires
    # both page_setup.fitToWidth/Height AND sheet_properties.pageSetUpPr,
    # otherwise Excel silently falls back to 100 % scale.
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 1.168
    ws.page_margins.right = 0.748
    ws.page_margins.top = 0.984
    ws.page_margins.bottom = 0.984
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0


def _build_header_block(
    ws: Worksheet,
    user: UserAccount,
    month_name: str,
    year: int,
) -> None:
    """Write rows 1–8: title, worker info, month/year labels."""
    for r in range(1, 9):
        _apply_row_height(ws, r)

    ws.merge_cells("A3:E3")
    _write_cell(
        ws,
        3,
        1,
        "Výkaz práce / odpracovaných hodin",
        font=_BOLD_FONT,
        fill=_CYAN_FILL,
        alignment=Alignment(horizontal="center"),
        border=_INFO_TITLE_A,
    )
    for col in (2, 3, 4):
        _write_cell(ws, 3, col, None, border=_INFO_TITLE_MID)
    _write_cell(ws, 3, 5, None, border=_INFO_TITLE_E)

    for row_num, label, value in (
        (4, "Jméno pracovníka:", user.name),
        (5, "Pracovní úvazek:", "DPP"),
        (6, "pozice:", "zdravotní dozory"),
    ):
        ws.merge_cells(f"A{row_num}:C{row_num}")
        ws.merge_cells(f"D{row_num}:E{row_num}")
        _write_cell(ws, row_num, 1, label, font=_STD_FONT, fill=_CYAN_FILL, border=_INFO_LABEL)
        _write_cell(ws, row_num, 2, None, border=_INFO_MID_B)
        _write_cell(ws, row_num, 3, None, border=_INFO_MID_C)
        _write_cell(ws, row_num, 4, value, font=_STD_FONT, border=_INFO_VALUE_D)
        _write_cell(ws, row_num, 5, None, border=_INFO_VALUE_E)

    _write_cell(ws, 7, 1, None, border=_INFO_SPACER_A)
    _write_cell(ws, 7, 2, None, border=_INFO_SPACER_B)
    _write_cell(ws, 7, 3, None, border=_INFO_SPACER_C)
    ws.merge_cells("D7:E7")
    _write_cell(ws, 7, 4, None, border=_INFO_SPACER_D)
    _write_cell(ws, 7, 5, None, border=_INFO_SPACER_E)

    _write_cell(ws, 8, 1, "Měsíc:", font=_STD_FONT, fill=_CYAN_FILL, border=_INFO_MONTH_A)
    _write_cell(ws, 8, 2, None, fill=_CYAN_FILL, border=_INFO_MONTH_B)
    _write_cell(ws, 8, 3, month_name, font=_STD_FONT, alignment=Alignment(horizontal="left"), border=_INFO_MONTH_C)
    _write_cell(ws, 8, 4, "Rok:", font=_STD_FONT, fill=_CYAN_FILL, border=_INFO_MONTH_D)
    _write_cell(ws, 8, 5, year, font=_STD_FONT, alignment=Alignment(horizontal="left"), border=_INFO_MONTH_E)


def _build_column_headers(ws: Worksheet) -> None:
    """Write the column-header row (row 9)."""
    _apply_row_height(ws, _HEADER_ROW)
    ws.merge_cells("D9:E9")
    _write_cell(
        ws,
        9,
        1,
        "Datum",
        font=_BOLD_FONT,
        fill=_BLUE_FILL,
        alignment=Alignment(horizontal="center"),
        border=_HDR_BORDER_A,
    )
    _write_cell(
        ws,
        9,
        2,
        "Den",
        font=_BOLD_FONT,
        fill=_BLUE_FILL,
        alignment=Alignment(horizontal="center"),
        border=_HDR_BORDER_B,
    )
    _write_cell(
        ws,
        9,
        3,
        "Počet hodin",
        font=_BOLD_FONT,
        fill=_BLUE_FILL,
        alignment=Alignment(horizontal="center", wrap_text=True),
        border=_HDR_BORDER_C,
    )
    _write_cell(
        ws,
        9,
        4,
        "Popis činnosti",
        font=_BOLD_FONT,
        fill=_BLUE_FILL,
        alignment=Alignment(horizontal="center"),
        border=_HDR_BORDER_D,
    )
    _write_cell(ws, 9, 5, None, border=_HDR_BORDER_E)


def _build_day_rows(
    ws: Worksheet,
    year: int,
    month: int,
    days_in_month: int,
    cz_holidays: set[date],
    events_by_day: dict[int, tuple[Decimal, list[str]]],
) -> None:
    """Write one row per calendar day (rows 10 … 10+days_in_month-1)."""
    for day in range(1, days_in_month + 1):
        row = _FIRST_DATA_ROW + day - 1
        _apply_row_height(ws, row)

        d = date(year, month, day)
        weekday = d.weekday()
        is_weekend = weekday >= 5
        is_holiday = d in cz_holidays

        fill = _YELLOW_FILL if is_holiday else _WHITE_FILL
        day_font = _RED_FONT if is_weekend else _STD_FONT

        hours_val, names = events_by_day.get(day, (None, []))
        hours_display = float(hours_val) if hours_val else None
        description = ", ".join(names) if names else None

        ws.merge_cells(f"D{row}:E{row}")
        _write_cell(
            ws, row, 1, day, font=_STD_FONT, fill=fill, alignment=Alignment(horizontal="center"), border=_DAY_BORDER_A
        )
        _write_cell(
            ws,
            row,
            2,
            CZ_WEEKDAY_ABBR[weekday],
            font=day_font,
            alignment=Alignment(horizontal="center"),
            border=_DAY_BORDER_B,
        )
        _write_cell(
            ws, row, 3, hours_display, font=_STD_FONT, alignment=Alignment(horizontal="center"), border=_DAY_BORDER_C
        )
        _write_cell(
            ws, row, 4, description, font=_STD_FONT, alignment=Alignment(horizontal="left"), border=_DAY_BORDER_D
        )
        _write_cell(ws, row, 5, None, border=_DAY_BORDER_E)


def _build_totals_and_signatures(
    ws: Worksheet,
    year: int,
    month: int,
    days_in_month: int,
    events_by_day: dict[int, tuple[Decimal, list[str]]],
    signature_image: bytes | None = None,
) -> None:
    """Write the totals row and signature rows below the day grid."""
    total_row = _FIRST_DATA_ROW + days_in_month
    _apply_row_height(ws, total_row)
    total_hours = float(sum(h for h, _ in events_by_day.values())) if events_by_day else 0.0
    ws.merge_cells(f"D{total_row}:E{total_row}")
    _write_cell(ws, total_row, 1, "Celkem hodin", font=_BOLD_FONT, border=_TOTAL_BORDER_A)
    _write_cell(ws, total_row, 2, None, font=_BOLD_FONT, border=_TOTAL_BORDER_B)
    _write_cell(
        ws, total_row, 3, total_hours, font=_BOLD_FONT, alignment=Alignment(horizontal="center"), border=_TOTAL_BORDER_C
    )
    _write_cell(ws, total_row, 4, None, border=_TOTAL_BORDER_D)
    _write_cell(ws, total_row, 5, None, border=_TOTAL_BORDER_E)

    sig_worker = total_row + 4
    sig_boss = total_row + 7
    for r in (sig_worker, sig_boss):
        _apply_row_height(ws, r)

    _write_cell(ws, sig_worker, 1, "Datum a podpis pracovníka:", font=_BOLD_FONT)
    last_day_date = date(year, month, days_in_month)
    _write_cell(
        ws,
        sig_worker,
        4,
        last_day_date,
        font=_STD_FONT,
        alignment=Alignment(horizontal="left"),
        number_format="DD.MM.YYYY",
    )
    _write_cell(ws, sig_boss, 1, "Datum a podpis nadřízeného pracovníka:", font=_BOLD_FONT)

    if signature_image:
        _embed_signature(ws, signature_image, sig_worker)


def _col_width_to_pixels(width_chars: float) -> int:
    """Approximate Excel column width in characters to pixels for Calibri 11.

    Uses the widely cited Excel formula (MDW=7): px = round(w*7 + 5) for w >= 1.
    Good enough for image placement; users can nudge manually if needed.
    """
    if width_chars < 1:
        return round(12 * width_chars)
    return round(width_chars * 7 + 5)


def _embed_signature(ws: Worksheet, signature_image: bytes, sig_row: int) -> None:
    """Anchor the user's signature image at the bottom-right of the signature row.

    Anchored so:
      - the bottom of the image sits on the bottom edge of the signature row;
      - the right edge is ~0.3 cm past the right edge of the last
        _COL_WIDTHS column (nudged rightward so that column's border is
        clearly visible next to it).
    Users can still drag the image after opening the xlsx for fine-tuning.
    """
    pil = PILImage.open(io.BytesIO(signature_image))
    ratio = _SIGNATURE_TARGET_HEIGHT_PX / pil.height
    img_w_px = max(1, round(pil.width * ratio))

    # Horizontal target: right edge = (sum of _COL_WIDTHS columns) + nudge.
    col_widths_emu = [pixels_to_EMU(_col_width_to_pixels(w)) for w in _COL_WIDTHS.values()]
    right_edge_emu = sum(col_widths_emu) + cm_to_EMU(_SIGNATURE_RIGHT_NUDGE_CM)
    left_emu = max(0, right_edge_emu - pixels_to_EMU(img_w_px))

    # Walk columns (extending beyond E with default-width columns) to find the
    # anchor cell and offset within it. Column indices in AnchorMarker are
    # 0-based.
    default_col_width_emu = pixels_to_EMU(_col_width_to_pixels(8.43))
    running = 0
    col_idx = 0
    while True:
        w = col_widths_emu[col_idx] if col_idx < len(col_widths_emu) else default_col_width_emu
        if left_emu < running + w:
            break
        running += w
        col_idx += 1
    offset_within_col_emu = left_emu - running

    # Vertical target: bottom of image on bottom edge of signature row.
    # Row height in points; 1 pt = 1/72 inch; EMU per inch = 914400.
    row_height_emu = round(_SIGNATURE_ROW_HEIGHT_PT * 914400 / 72)
    img_h_emu = pixels_to_EMU(_SIGNATURE_TARGET_HEIGHT_PX)
    row_off_emu = max(0, row_height_emu - img_h_emu)

    img = XlsxImage(io.BytesIO(signature_image))
    marker = AnchorMarker(
        col=col_idx,
        colOff=offset_within_col_emu,
        row=sig_row - 1,
        rowOff=row_off_emu,
    )
    ext = XDRPositiveSize2D(
        cx=pixels_to_EMU(img_w_px),
        cy=img_h_emu,
    )
    img.anchor = OneCellAnchor(_from=marker, ext=ext)

    ws.row_dimensions[sig_row].height = _SIGNATURE_ROW_HEIGHT_PT
    ws.add_image(img)


# ── Core generator ────────────────────────────────────────────────────────────


def generate_work_report(user: UserAccount, year: int, month: int) -> Path:
    """
    Build the výkaz práce xlsx for *user* for the given *year*/*month*.

    The file is written to  instance/work_report/<user_id>/<year>-<MM>.xlsx
    and is overwritten if it already exists.  Returns the absolute Path.
    """
    month_name = CZ_MONTH_NAMES[month]
    days_in_month = calendar.monthrange(year, month)[1]
    cz_holidays: set[date] = set(holidays.CZ(years=year).keys())
    events_by_day = _fetch_events_for_month(str(user.id), year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = month_name

    _setup_worksheet(ws, month_name)
    _build_header_block(ws, user, month_name, year)
    _build_column_headers(ws)
    _build_day_rows(ws, year, month, days_in_month, cz_holidays, events_by_day)
    _build_totals_and_signatures(ws, year, month, days_in_month, events_by_day, signature_image=user.signature_image)

    # Scope fit-to-page to the report block; without this Excel would try to
    # fit any incidentally-referenced empty rows too, shrinking the output.
    last_row = _FIRST_DATA_ROW + days_in_month + 7  # totals + 2 signature rows
    ws.print_area = f"A1:E{last_row}"

    instance_path = Path(current_app.instance_path)
    out_dir = instance_path / "work_report" / str(user.id)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{year}-{month:02d}.xlsx"
    wb.save(str(out_path))
    return out_path
