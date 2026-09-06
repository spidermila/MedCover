"""
Shared Excel primitives — styles, cell writing and a generic table-sheet builder.

Reports that need an xlsx download describe their output as a list of
:class:`TableSheet` objects and hand them to :func:`build_workbook`::

    wb = build_workbook([
        TableSheet(
            sheet_name="Výkazy",
            title="Přehled výkazů",
            subtitle="Období: 2026-01-01 – 2026-01-31",
            columns=[Column("Jméno", 26), Column("Hodiny", 12, align="right", number_format=HOURS_FORMAT)],
            rows=[["Jan Novák", Decimal("7.5")]],
        )
    ])

``Decimal`` values are written as real numbers so Excel sums them regardless
of the user's decimal separator; ``date``/``datetime`` values become real date
cells. Everything else is written as-is, with leading formula characters
neutralised.
"""

from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="FF4472C4")  # dark blue
TITLE_FILL = PatternFill("solid", fgColor="FFD9E1F2")  # light blue

HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
TITLE_FONT = Font(name="Calibri", size=10, bold=True)
INFO_FONT = Font(name="Calibri", size=9)
STD_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)


def _side(style: str) -> Side:
    return Side(border_style=style)


THIN = Border(
    left=_side("thin"),
    right=_side("thin"),
    top=_side("thin"),
    bottom=_side("thin"),
)

# Number formats
HOURS_FORMAT = "0.0"
DATE_FORMAT = "DD.MM.YYYY"

# A cell whose text starts with one of these is interpreted as a formula by
# Excel, so user-supplied values are prefixed with an apostrophe first.
FORMULA_STARTERS = ("=", "+", "-", "@", "\t", "\r")


# ── Primitives ────────────────────────────────────────────────────────────────


def cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: object = None,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    """Write a single styled cell, neutralising values Excel would read as formulas."""
    if isinstance(value, str) and value.startswith(FORMULA_STARTERS):
        value = "'" + value
    target = ws.cell(row=row, column=col, value=value)
    if font is not None:
        target.font = font
    if fill is not None:
        target.fill = fill
    if alignment is not None:
        target.alignment = alignment
    if border is not None:
        target.border = border
    if number_format is not None:
        target.number_format = number_format


def title_block(ws: Worksheet, title: str, subtitle: str, n_cols: int) -> int:
    """Write a two-row title block merged across all columns. Returns the next row number."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell(ws, 1, 1, title, font=TITLE_FONT, fill=TITLE_FILL, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    cell(ws, 2, 1, subtitle, font=INFO_FONT, fill=TITLE_FILL, alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 15

    return 3  # header row


# ── Generic table sheet ───────────────────────────────────────────────────────


@dataclass(frozen=True)
class Column:
    """One column of a table sheet."""

    header: str
    width: int
    align: str = "left"
    number_format: str | None = None

    @property
    def alignment(self) -> Alignment:
        return Alignment(horizontal=self.align, vertical="center")


@dataclass(frozen=True)
class TableSheet:
    """A single worksheet: title block, header row, data rows, optional totals row."""

    sheet_name: str
    title: str
    subtitle: str
    columns: list[Column]
    rows: list[Sequence[object]]
    totals_row: Sequence[object] | None = None
    autofilter: bool = True


def _write_row(ws: Worksheet, row: int, columns: list[Column], values: Sequence[object], font: Font) -> None:
    for i, (column, value) in enumerate(zip(columns, values, strict=True), 1):
        number_format = column.number_format
        if isinstance(value, Decimal):
            # Real numbers, not preformatted strings, so Excel sums them
            # whatever the viewer's decimal separator is.
            value = float(value)
        if isinstance(value, date) and number_format is None:
            number_format = DATE_FORMAT
        cell(
            ws,
            row,
            i,
            value,
            font=font,
            alignment=column.alignment,
            border=THIN,
            number_format=number_format,
        )


def build_table_sheet(ws: Worksheet, sheet: TableSheet) -> None:
    """Render *sheet* onto *ws*."""
    n_cols = len(sheet.columns)
    header_row = title_block(ws, sheet.title, sheet.subtitle, n_cols)

    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, column in enumerate(sheet.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = column.width
        cell(ws, header_row, i, column.header, font=HEADER_FONT, fill=HEADER_FILL, alignment=centre, border=THIN)
    ws.row_dimensions[header_row].height = 30

    row = header_row + 1
    for values in sheet.rows:
        _write_row(ws, row, sheet.columns, values, STD_FONT)
        row += 1

    # The autofilter must not cover the totals row, or sorting would move it.
    if sheet.autofilter:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{max(row - 1, header_row)}"

    if sheet.totals_row is not None:
        _write_row(ws, row, sheet.columns, sheet.totals_row, BOLD_FONT)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def build_workbook(sheets: list[TableSheet]) -> Workbook:
    """Build a workbook from one or more table sheets. Caller saves/streams it."""
    wb = Workbook()
    first = wb.active
    first.title = sheets[0].sheet_name
    build_table_sheet(first, sheets[0])
    for sheet in sheets[1:]:
        build_table_sheet(wb.create_sheet(title=sheet.sheet_name), sheet)
    return wb
