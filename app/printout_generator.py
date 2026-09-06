"""
Printout generator — Excel export for physical event overviews.

Generates a workbook with two sheets:
  - Podpisy: one row per (event, spot), wide signature column for handwriting
  - Přehled: one row per event, spots side-by-side in columns

Public entry point::

    wb = generate_printout(events, date_range, me_name)
    # caller saves to a BytesIO buffer and serves as a download
"""

from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils import get_app_tz
from app.xlsx import HEADER_FILL, HEADER_FONT, STD_FONT, THIN, cell, title_block

if TYPE_CHECKING:
    from app.models.event import Event


# ── Sheet 1: Podpisy (Signatures) ─────────────────────────────────────────────


def _build_signature_sheet(
    ws: Worksheet,
    events: list[Event],
    date_range: str,
    me_name: str | None,
) -> None:
    headers = ["Datum", "Název akce", "Jméno", "Kvalifikace", "Popis pozice", "Podpis"]
    widths = [12, 34, 26, 22, 22, 38]

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    n_cols = len(headers)
    subtitle = _subtitle(date_range, me_name)
    hdr_row = title_block(ws, "Sestava pro tisk — Podpisy", subtitle, n_cols)

    # Column headers
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, h in enumerate(headers, 1):
        cell(ws, hdr_row, i, h, font=HEADER_FONT, fill=HEADER_FILL, alignment=centre, border=THIN)
    ws.row_dimensions[hdr_row].height = 18

    # Data — one row per (event, spot)
    tz = get_app_tz()
    row = hdr_row + 1
    left = Alignment(horizontal="left", vertical="center")

    for event in events:
        spots = sorted(event.spots, key=lambda s: s.id)
        if not spots:
            continue
        date_str = event.start_datetime.astimezone(tz).strftime("%d.%m.%Y")

        for spot in spots:
            person = spot.assignment.user.name if spot.assignment else ""
            quals = ", ".join(q.name for q in spot.required_qualifications if not q.is_deleted)
            desc = spot.description or ""

            for col, val in enumerate([date_str, event.name, person, quals, desc, ""], 1):
                cell(ws, row, col, val, font=STD_FONT, alignment=left, border=THIN)

            ws.row_dimensions[row].height = 28  # room for handwriting
            row += 1


# ── Sheet 2: Přehled (Overview) ───────────────────────────────────────────────


def _build_overview_sheet(
    ws: Worksheet,
    events: list[Event],
    date_range: str,
    me_name: str | None,
) -> None:
    max_spots = max((len(e.spots) for e in events), default=1)

    fixed_headers = ["Datum", "Název akce", "Stav"]
    fixed_widths = [12, 34, 18]
    spot_width = 28

    for i, w in enumerate(fixed_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for j in range(max_spots):
        ws.column_dimensions[get_column_letter(len(fixed_headers) + j + 1)].width = spot_width

    n_cols = len(fixed_headers) + max_spots
    subtitle = _subtitle(date_range, me_name)
    hdr_row = title_block(ws, "Sestava pro tisk — Přehled", subtitle, n_cols)

    # Column headers
    centre = Alignment(horizontal="center", vertical="center")
    for i, h in enumerate(fixed_headers, 1):
        cell(ws, hdr_row, i, h, font=HEADER_FONT, fill=HEADER_FILL, alignment=centre, border=THIN)
    for j in range(max_spots):
        cell(
            ws,
            hdr_row,
            len(fixed_headers) + j + 1,
            f"Pozice {j + 1}",
            font=HEADER_FONT,
            fill=HEADER_FILL,
            alignment=centre,
            border=THIN,
        )
    ws.row_dimensions[hdr_row].height = 18

    # Data — one row per event
    tz = get_app_tz()
    row = hdr_row + 1
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for event in events:
        spots = sorted(event.spots, key=lambda s: s.id)
        date_str = event.start_datetime.astimezone(tz).strftime("%d.%m.%Y")

        for col, val in enumerate([date_str, event.name, event.status.value], 1):
            cell(
                ws,
                row,
                col,
                val,
                font=STD_FONT,
                alignment=Alignment(horizontal="left", vertical="center"),
                border=THIN,
            )

        for j in range(max_spots):
            if j < len(spots):
                spot = spots[j]
                cell_val = spot.assignment.user.name if spot.assignment else ""
            else:
                cell_val = ""
            cell(ws, row, len(fixed_headers) + j + 1, cell_val, font=STD_FONT, alignment=left, border=THIN)

        ws.row_dimensions[row].height = 18
        row += 1


# ── Public entry point ────────────────────────────────────────────────────────


def _subtitle(date_range: str, me_name: str | None) -> str:
    s = f"Období: {date_range}"
    if me_name:
        s += f"  |  Nadřazená akce: {me_name}"
    return s


def generate_printout(
    events: list[Event],
    date_range: str,
    me_name: str | None,
) -> Workbook:
    """Build the printout workbook. Caller is responsible for saving/streaming."""
    wb = Workbook()

    ws_sig = wb.active
    ws_sig.title = "Podpisy"
    _build_signature_sheet(ws_sig, events, date_range, me_name)

    ws_overview = wb.create_sheet(title="Přehled")
    _build_overview_sheet(ws_overview, events, date_range, me_name)

    return wb
